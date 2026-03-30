"""Quick script to verify per-user cost allocation accuracy."""
import openpyxl

wb = openpyxl.load_workbook("Microsoft_License_Allocation_Model_March_2026.xlsx")

# SKU_Cost_Input
sci = wb["SKU_Cost_Input"]
skus = []
for r in range(2, 20):
    name = sci.cell(r, 1).value
    if not name or name == "TOTAL":
        break
    amt = sci.cell(r, 4).value
    amount = float(amt) if isinstance(amt, (int, float)) else 0.0
    active = sci.cell(r, 2).value or 0
    active = int(active) if isinstance(active, (int, float)) else 0
    skus.append({"name": name, "active_users": active, "amount": amount})

# License_Raw
lr = wb["License_Raw"]
license_map = {}
for r in range(2, lr.max_row + 1):
    display = lr.cell(r, 1).value
    email = lr.cell(r, 2).value
    sku_str = lr.cell(r, 3).value or ""
    if not display:
        break
    for sku in str(sku_str).split("+"):
        sku = sku.strip()
        if sku:
            license_map.setdefault((email or "").lower(), []).append(sku)

# Unit costs with penny allocation
precise_unit = {}
sku_remainder = {}
for s in skus:
    if s["active_users"] > 0 and s["amount"] > 0:
        unit = s["amount"] / s["active_users"]
        rounded_unit = round(unit, 2)
        remainder = round(s["amount"] - (rounded_unit * s["active_users"]), 2)
        precise_unit[s["name"]] = rounded_unit
        sku_remainder[s["name"]] = remainder
    else:
        precise_unit[s["name"]] = 0.0

# User_Match
um = wb["User_Match"]
users = []
for r in range(2, um.max_row + 1):
    name = um.cell(r, 1).value
    if not name:
        break
    users.append({
        "email": (um.cell(r, 2).value or "").lower(),
        "display_name": name,
        "division": um.cell(r, 6).value or "",
    })

sku_user_totals = {}
for u in users:
    for s in license_map.get(u["email"], []):
        sku_user_totals[s] = sku_user_totals.get(s, 0) + 1

sku_seen = {}
total_all = 0
per_sku_allocated = {}
for u in users:
    user_skus = license_map.get(u["email"], [])
    total_cost = 0.0
    for s in user_skus:
        sku_seen[s] = sku_seen.get(s, 0) + 1
        cost = precise_unit.get(s, 0)
        if sku_seen[s] == sku_user_totals.get(s, 0) and s in sku_remainder:
            cost += sku_remainder[s]
        total_cost += cost
        per_sku_allocated[s] = per_sku_allocated.get(s, 0) + cost
    u["allocated_cost"] = round(total_cost, 2)
    total_all += u["allocated_cost"]

print("=== Per-SKU allocation totals vs invoice amounts ===")
for s in skus:
    if s["amount"] > 0:
        alloc = round(per_sku_allocated.get(s["name"], 0), 2)
        diff = round(alloc - s["amount"], 2)
        print(f"  {s['name']}: Invoice R{s['amount']:,.2f} | Allocated R{alloc:,.2f} | Diff R{diff:,.2f}")

print(f"\nTotal allocated: R{total_all:,.2f}")
print(f"Invoice total:   R52,334.23")
print(f"Gap:             R{52334.23 - total_all:,.2f} (= Power Apps unallocable)")

# Invoice unit prices for comparison
invoice_unit_prices = {
    "Microsoft 365 Business Standard": 209.55,
    "Microsoft 365 E3": 603.29,
    "Microsoft 365 Business Premium": 368.68,
    "Power BI Premium Per User": 402.19,
    "Exchange Online (Plan 1)": 67.03,
    "Power Automate per user plan": 251.37,
    "Microsoft Defender for Office 365 (Plan 2)": 83.79,
    "Power Apps per app plan (1 app or website)": 83.79,
}

print("\n=== Unit cost comparison: Model (effective) vs Invoice (list price) ===")
for s in skus:
    if s["amount"] > 0:
        model_unit = precise_unit.get(s["name"], 0)
        inv_unit = invoice_unit_prices.get(s["name"], 0)
        diff = round(model_unit - inv_unit, 2)
        status = "MATCH" if diff == 0 else f"DIFF R{diff:+.2f}"
        print(f"  {s['name']}:")
        print(f"    Model effective: R{model_unit:.2f}  |  Invoice list: R{inv_unit:.2f}  |  {status}")

# Show sample users with detailed cost breakdown
print("\n=== Sample user costs (first 15) ===")
for u in users[:15]:
    user_skus = license_map.get(u["email"], [])
    sku_costs = []
    for s in user_skus:
        sku_costs.append(f"{s}: R{precise_unit.get(s, 0):.2f}")
    print(f"  {u['display_name']} ({u['division']}): R{u['allocated_cost']:.2f}")
    if sku_costs:
        print(f"    = {' + '.join(sku_costs)}")
