"""
Build Microsoft License Cost Allocation Model - March 2026
Reads source data, builds a clean Excel workbook with all required sheets,
formulas, formatting, and audit trail.
Includes: Invoice reconciliation, VAT handling, exchange rate, proration policy.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy
import os

# === Paths ===
BASE = r"c:\Users\nadhi\OneDrive - Wearcheck Reliability Solutions\Desktop\Microsoft Monthly Licensing"
SRC = os.path.join(BASE, "Source Data")
EMP_FILE = os.path.join(SRC, "employee_master_list.xlsx")
LIC_FILE = os.path.join(SRC, "Microsoft_License_Allocation_March_2026.xlsx")
OUTPUT = os.path.join(BASE, "Microsoft_License_Allocation_Model_March_2026.xlsx")

# === Styles ===
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
KPI_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GOOD_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ERR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
RS_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
AFS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NAM_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
MOZ_FILL = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CURRENCY_FMT = '#,##0.00'
PCT_FMT = '0.0%'

DIVISIONS = ["RS", "AFS", "Namibia", "Mozambique"]
DIV_FILLS = {"RS": RS_FILL, "AFS": AFS_FILL, "Namibia": NAM_FILL, "Mozambique": MOZ_FILL, "Unassigned - Review": REVIEW_FILL}

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_data_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")

def auto_width(ws, max_col, max_width=40):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            val = row[0]
            if val is not None:
                s = str(val)
                if s.startswith("="):
                    s = "X" * 12  # estimate formula width
                max_len = max(max_len, len(s))
        ws.column_dimensions[letter].width = min(max_len + 3, max_width)

def freeze_and_filter(ws, freeze_cell, max_col, header_row):
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{ws.max_row}"

# === Load source data ===
print("Loading source data...")
lic_wb = openpyxl.load_workbook(LIC_FILE)

# --- Employee Master ---
emp_wb = openpyxl.load_workbook(EMP_FILE)
emp_ws = emp_wb["Employees Details (Onboarding a"]
emp_headers = [c.value for c in list(emp_ws.iter_rows(min_row=1, max_row=1))[0]]

# Find key columns
def find_col(headers, name):
    for i, h in enumerate(headers):
        if h and name.lower() in str(h).lower():
            return i
    return None

email_idx = find_col(emp_headers, "Company E-Mail")
fname_idx = find_col(emp_headers, "First Names")
sname_idx = find_col(emp_headers, "Surname")
code_idx = find_col(emp_headers, "Employee Code")
dept1_idx = find_col(emp_headers, "Department")  # first Department column
dept2_idx = None
# Find second Department column
dept_count = 0
for i, h in enumerate(emp_headers):
    if h and str(h).strip().lower() == "department":
        dept_count += 1
        if dept_count == 1:
            dept1_idx = i
        elif dept_count == 2:
            dept2_idx = i

site_idx = find_col(emp_headers, "Site")
title_idx = find_col(emp_headers, "Job Title")
super_idx = find_col(emp_headers, "Direct Supervisor")
cell_idx = find_col(emp_headers, "Cell Number")
status_idx = find_col(emp_headers, "Status")

print(f"Employee columns: email={email_idx}, fname={fname_idx}, sname={sname_idx}, code={code_idx}, dept1={dept1_idx}, dept2={dept2_idx}")

# Build employee records
employees = []
for row in emp_ws.iter_rows(min_row=2, max_row=emp_ws.max_row, values_only=True):
    row = list(row)
    email = row[email_idx] if email_idx is not None and email_idx < len(row) else None
    if not email:
        continue
    fname = row[fname_idx] if fname_idx is not None and fname_idx < len(row) else ""
    sname = row[sname_idx] if sname_idx is not None and sname_idx < len(row) else ""
    code = row[code_idx] if code_idx is not None and code_idx < len(row) else ""
    dept1 = row[dept1_idx] if dept1_idx is not None and dept1_idx < len(row) else ""
    dept2 = row[dept2_idx] if dept2_idx is not None and dept2_idx < len(row) else ""
    site = row[site_idx] if site_idx is not None and site_idx < len(row) else ""
    title = row[title_idx] if title_idx is not None and title_idx < len(row) else ""
    supervisor = row[super_idx] if super_idx is not None and super_idx < len(row) else ""
    cell_num = row[cell_idx] if cell_idx is not None and cell_idx < len(row) else ""
    status = row[status_idx] if status_idx is not None and status_idx < len(row) else ""

    name = f"{fname} {sname}".strip() if fname and sname else (fname or sname or "")

    # Derive division
    dept2_str = str(dept2).strip() if dept2 else ""
    dept1_str = str(dept1).strip() if dept1 else ""
    site_str = str(site).strip() if site else ""

    if dept2_str == "AFS" or dept1_str == "AFS":
        division = "AFS"
    elif "namibia" in dept2_str.lower() or "namibia" in dept1_str.lower() or "namibia" in site_str.lower():
        division = "Namibia"
    elif "mozambique" in dept2_str.lower() or "mozambique" in dept1_str.lower() or "mozambique" in site_str.lower() or "moz" in dept2_str.lower():
        division = "Mozambique"
    elif dept2_str in ("GP Consult", "WearCheck"):
        division = "RS"
    else:
        division = "RS"  # Default to RS

    employees.append({
        "email": str(email).strip(),
        "email_lower": str(email).strip().lower(),
        "name": name,
        "code": str(code).strip() if code else "",
        "dept1": dept1_str,
        "dept2": dept2_str,
        "site": str(site).strip() if site else "",
        "division": division,
        "title": str(title).strip() if title else "",
        "supervisor": str(supervisor).strip() if supervisor else "",
        "cell": str(cell_num).strip() if cell_num else "",
        "status": str(status).strip() if status else "",
    })

# Build email lookup (case-insensitive)
emp_by_email = {}
for e in employees:
    emp_by_email[e["email_lower"]] = e

print(f"Loaded {len(employees)} employees")

# --- License Raw ---
lic_raw_ws = lic_wb["License_Raw"]
licensed_users = []
for row in lic_raw_ws.iter_rows(min_row=2, max_row=lic_raw_ws.max_row, values_only=True):
    display_name, email, licenses = row[0], row[1], row[2]
    if not email:
        continue
    sku_list = [s.strip() for s in str(licenses).split("+") if s.strip()] if licenses else []
    licensed_users.append({
        "display_name": str(display_name).strip() if display_name else "",
        "email": str(email).strip(),
        "email_lower": str(email).strip().lower(),
        "skus": sku_list,
    })

print(f"Loaded {len(licensed_users)} licensed users")

# --- Collect all unique SKUs ---
all_skus = set()
for u in licensed_users:
    all_skus.update(u["skus"])
all_skus = sorted(all_skus)
print(f"Found {len(all_skus)} unique SKUs")

# --- Count users per SKU ---
sku_user_counts = {}
for sku in all_skus:
    count = sum(1 for u in licensed_users if sku in u["skus"])
    sku_user_counts[sku] = count

# --- Also load existing User_Match data for higher-quality matching ---
existing_match_ws = lic_wb["User_Match"]
existing_matches = {}
for row in existing_match_ws.iter_rows(min_row=2, max_row=existing_match_ws.max_row, values_only=True):
    row = list(row)
    lic_email = str(row[1]).strip().lower() if row[1] else ""
    matched_name = row[2]
    candidate_email = str(row[3]).strip() if row[3] else ""
    emp_number = row[4]
    division = row[5]
    match_status = row[9]
    confidence = row[10]
    review_req = row[11]
    mapping_note = row[12]
    existing_matches[lic_email] = {
        "matched_name": matched_name,
        "candidate_email": candidate_email,
        "emp_number": emp_number,
        "division": division,
        "match_status": match_status,
        "confidence": confidence,
        "review_required": review_req,
        "mapping_note": mapping_note,
    }

# --- Manual overrides from review (March 2026) ---
# Format: email_lower → {division, match_status, confidence, review_required, mapping_note, user_status}
# user_status: "Active", "Ex-Employee", "Contractor", "Service Account", "Management Review"
MANUAL_OVERRIDES = {
    # Confirmed matches
    "jean-pierre@wearcheckrs.com": {
        "division": "AFS", "match_status": "Confirmed", "confidence": "High",
        "review_required": "No", "mapping_note": "Confirmed match via exact name (reviewed)",
        "user_status": "Active",
    },
    "joe@wearcheckrs.com": {
        "division": "AFS", "match_status": "Confirmed", "confidence": "High",
        "review_required": "No", "mapping_note": "Confirmed match via manual alias/name rule (reviewed)",
        "user_status": "Active",
    },
    "johan.stols@wearcheckrs.com": {
        "division": "AFS", "match_status": "Confirmed", "confidence": "High",
        "review_required": "No", "mapping_note": "Confirmed match via manual alias/name rule (reviewed)",
        "user_status": "Active",
    },
    "lorraine@wearcheckrs.com": {
        "division": "AFS", "match_status": "Confirmed", "confidence": "High",
        "review_required": "No", "mapping_note": "Confirmed: she belongs here (reviewed)",
        "user_status": "Active",
    },
    "mervyn@wearcheckrs.com": {
        "division": "AFS", "match_status": "Confirmed", "confidence": "High",
        "review_required": "No", "mapping_note": "Confirmed match to mervyng@wearcheck.co.za (reviewed)",
        "user_status": "Active",
    },
    "christene@wearcheckrs.com": {
        "division": "RS", "match_status": "Confirmed", "confidence": "High",
        "review_required": "No", "mapping_note": "Confirmed email alias match to chrstene@wearcheckrs.com (reviewed)",
        "user_status": "Active",
    },
    # Newly added to employee list — will re-match via email
    "boitumelo@wearcheckrs.com": {
        "division": "RS", "match_status": "Confirmed", "confidence": "High",
        "review_required": "No", "mapping_note": "User now added to employee master (reviewed)",
        "user_status": "Active",
    },
    "stephanied@wearcheckrs.com": {
        "division": "RS", "match_status": "Confirmed", "confidence": "High",
        "review_required": "No", "mapping_note": "Missed on employee list — now added (reviewed)",
        "user_status": "Active",
    },
    # Old/duplicate account
    "adri.ludick@wearcheckrs.com": {
        "division": "RS", "match_status": "Ex-Employee", "confidence": "High",
        "review_required": "No",
        "mapping_note": "Old account — new email A.Ludick@WearCheckRS.com is active. FLAG: Remove this licence.",
        "user_status": "Ex-Employee",
    },
    # No longer employees — FLAG FOR LICENCE REMOVAL
    "christian@wearcheckrs.com": {
        "division": "RS", "match_status": "Ex-Employee", "confidence": "High",
        "review_required": "No", "mapping_note": "No longer an employee. FLAG: Remove licence.",
        "user_status": "Ex-Employee",
    },
    "gladwins@wearcheckrs.com": {
        "division": "RS", "match_status": "Ex-Employee", "confidence": "High",
        "review_required": "No", "mapping_note": "No longer an employee. FLAG: Remove licence.",
        "user_status": "Ex-Employee",
    },
    "ivan@wearcheckrs.com": {
        "division": "RS", "match_status": "Ex-Employee", "confidence": "High",
        "review_required": "No", "mapping_note": "No longer an employee. FLAG: Remove licence.",
        "user_status": "Ex-Employee",
    },
    "kenneth@wearcheckrs.com": {
        "division": "RS", "match_status": "Ex-Employee", "confidence": "High",
        "review_required": "No", "mapping_note": "No longer an employee. FLAG: Remove licence.",
        "user_status": "Ex-Employee",
    },
    "kevinj@wearcheckrs.com": {
        "division": "RS", "match_status": "Ex-Employee", "confidence": "High",
        "review_required": "No", "mapping_note": "No longer an employee. FLAG: Remove licence.",
        "user_status": "Ex-Employee",
    },
    "mandiso@wearcheckrs.com": {
        "division": "RS", "match_status": "Ex-Employee", "confidence": "High",
        "review_required": "No", "mapping_note": "No longer an employee. FLAG: Remove licence.",
        "user_status": "Ex-Employee",
    },
    "martin.prinsloo@wearcheckrs.com": {
        "division": "RS", "match_status": "Ex-Employee", "confidence": "High",
        "review_required": "No", "mapping_note": "No longer an employee. FLAG: Remove licence.",
        "user_status": "Ex-Employee",
    },
    "peter@wearcheckrs.com": {
        "division": "RS", "match_status": "Ex-Employee", "confidence": "High",
        "review_required": "No", "mapping_note": "No longer an employee. FLAG: Remove licence.",
        "user_status": "Ex-Employee",
    },
    "thabang@wearcheckrs.com": {
        "division": "RS", "match_status": "Ex-Employee", "confidence": "High",
        "review_required": "No", "mapping_note": "No longer an employee. FLAG: Remove licence.",
        "user_status": "Ex-Employee",
    },
    "tumelo.samancor@wearcheckrs.com": {
        "division": "RS", "match_status": "Ex-Employee", "confidence": "High",
        "review_required": "No", "mapping_note": "No longer an employee. FLAG: Remove licence.",
        "user_status": "Ex-Employee",
    },
    # Contractors — keep licence, tag as contractor, allocate to RS
    "lea@wearcheckrs.com": {
        "division": "RS", "match_status": "Contractor", "confidence": "High",
        "review_required": "No", "mapping_note": "Not an employee — contractor. Licence allocated to RS.",
        "user_status": "Contractor",
    },
    "marthinus@wearcheckrs.com": {
        "division": "RS", "match_status": "Contractor", "confidence": "High",
        "review_required": "No", "mapping_note": "Not an employee — contractor. Licence allocated to RS.",
        "user_status": "Contractor",
    },
    "nicoliens@wearcheckrs.com": {
        "division": "RS", "match_status": "Contractor", "confidence": "High",
        "review_required": "No", "mapping_note": "Not an employee — contractor. Licence allocated to RS.",
        "user_status": "Contractor",
    },
    # Management review
    "christie.samancor@wearcheckrs.com": {
        "division": "RS", "match_status": "Management Review", "confidence": "Low",
        "review_required": "Yes",
        "mapping_note": "Not an employee — has free Fabric licence only. Flagged for management review.",
        "user_status": "Management Review",
    },
}

# --- Match licensed users to employees ---
def match_user(user):
    email_low = user["email_lower"]

    # 0. Manual override takes highest priority
    if email_low in MANUAL_OVERRIDES:
        ov = MANUAL_OVERRIDES[email_low]
        return {
            "matched_name": "",
            "candidate_email": "",
            "emp_number": "",
            "division": ov["division"],
            "match_status": ov["match_status"],
            "confidence": ov["confidence"],
            "review_required": ov["review_required"],
            "mapping_note": ov["mapping_note"],
            "user_status": ov["user_status"],
        }

    # 1. Direct email match against employee master
    if email_low in emp_by_email:
        emp = emp_by_email[email_low]
        return {
            "matched_name": emp["name"],
            "candidate_email": emp["email"],
            "emp_number": emp["code"],
            "division": emp["division"],
            "match_status": "Matched",
            "confidence": "High",
            "review_required": "No",
            "mapping_note": "Exact company email match",
            "user_status": "Active",
        }

    # 2. Use existing match from pre-processed file
    if email_low in existing_matches:
        em = existing_matches[email_low]
        div = em["division"]
        conf = em["confidence"]
        review = em["review_required"]
        status = em["match_status"]
        note = em["mapping_note"]

        # Apply RS default rule: unassigned users get RS temporarily
        if div == "Unassigned - Review" or not div:
            div = "RS"
            note = (note or "") + " | Auto-assigned to RS (unmapped user)"
            review = "Yes"
            status = "Auto-RS"
            conf = "Low"

        return {
            "matched_name": em["matched_name"],
            "candidate_email": em["candidate_email"],
            "emp_number": em["emp_number"],
            "division": div,
            "match_status": status,
            "confidence": conf,
            "review_required": review,
            "mapping_note": note,
            "user_status": "Active",
        }

    # 3. Service/shared/admin accounts → RS
    name_low = user["display_name"].lower()
    if any(kw in name_low for kw in ["admin", "shared", "service", "noreply", "test"]):
        return {
            "matched_name": "",
            "candidate_email": "",
            "emp_number": "",
            "division": "RS",
            "match_status": "Service Account",
            "confidence": "Rule",
            "review_required": "No",
            "mapping_note": "Service/shared/generic account → auto-assigned to RS",
            "user_status": "Service Account",
        }

    # 4. Unmatched → assign to RS with review flag
    return {
        "matched_name": "",
        "candidate_email": "",
        "emp_number": "",
        "division": "RS",
        "match_status": "Auto-RS",
        "confidence": "Low",
        "review_required": "Yes",
        "mapping_note": "No employee match found → temporarily assigned to RS",
        "user_status": "Unknown",
    }

matched_users = []
for u in licensed_users:
    m = match_user(u)
    matched_users.append({**u, **m})

print(f"Matched {sum(1 for m in matched_users if m['confidence'] == 'High')} high confidence")
print(f"Review required: {sum(1 for m in matched_users if m['review_required'] == 'Yes')}")

# --- Billing hints ---
def billing_hint(sku_name):
    name_low = sku_name.lower()
    if "free" in name_low or "trial" in name_low or "viral" in name_low:
        return "Likely free / trial"
    return "Likely billable"

# ===================================================
# BUILD OUTPUT WORKBOOK
# ===================================================
print("\nBuilding workbook...")
wb = openpyxl.Workbook()

# ===================================================
# Sheet 1: Read_Me
# ===================================================
ws = wb.active
ws.title = "Read_Me"
ws.sheet_properties.tabColor = "2F5496"

readme_lines = [
    ("Microsoft License Cost Allocation Model - March 2026", TITLE_FONT, None),
    ("", None, None),
    ("Purpose", SUBTITLE_FONT, None),
    ("Allocate total Microsoft license costs to internal cost centres (RS, AFS, Namibia, Mozambique)", None, None),
    ("based on active licensed users, using per-license SKU allocation aligned with Microsoft invoicing.", None, None),
    ("", None, None),
    ("How To Use (Monthly)", SUBTITLE_FONT, None),
    ("1. Complete Invoice_Input: enter invoice total (incl VAT), VAT rate, exchange rate, billing period", None, None),
    ("2. Replace License_Raw data with the new month's Microsoft Admin Centre export", None, None),
    ("3. Update Employee_Master if there are new starters/leavers", None, None),
    ("4. Enter monthly invoice line-item amounts per SKU on SKU_Cost_Input (yellow cells)", None, None),
    ("5. CHECK: Verify the Reconciliation section on Division_Summary — variance must be R0.00", None, None),
    ("6. Review the Review_Queue for any unresolved user mappings", None, None),
    ("7. Send Division_Summary to Finance", None, None),
    ("", None, None),
    ("Key Rules", SUBTITLE_FONT, None),
    ("- Email address is the unique key for matching users across datasets", None, None),
    ("- Cost per SKU ÷ number of users on that SKU = unit cost per user", None, None),
    ("- Allocation based on user's home division (cross-division work ignored)", None, None),
    ("- Shared/generic/service accounts auto-assigned to RS", None, None),
    ("- Unmatched users temporarily assigned to RS until confirmed (flagged for review)", None, None),
    ("- Allocation is based on users active at time of invoice", None, None),
    ("", None, None),
    ("Invoice & Pricing Rules", SUBTITLE_FONT, None),
    ("- SKU prices on SKU_Cost_Input must match the ACTUAL Microsoft invoice line items (ZAR, excl VAT)", None, None),
    ("- VAT is applied to the total via Invoice_Input — do NOT include VAT in individual SKU amounts", None, None),
    ("- If invoice is in USD, convert to ZAR using the exchange rate on Invoice_Input before entering", None, None),
    ("- The Reconciliation check ensures SKU total exactly matches the invoice — any difference is flagged red", None, None),
    ("", None, None),
    ("Proration & Snapshot Policy", SUBTITLE_FONT, None),
    ("- Allocation uses a snapshot: who is on the user list at the time of invoice processing", None, None),
    ("- If Microsoft prorates mid-month additions, they appear as fractional amounts on the invoice", None, None),
    ("- Any prorated delta that doesn't split evenly to a SKU → enter the exact invoice amount per SKU", None, None),
    ("- The model will self-correct via the reconciliation check (SKU totals must equal invoice total)", None, None),
    ("", None, None),
    ("Cost Centres", SUBTITLE_FONT, None),
    ("RS, AFS, Namibia, Mozambique", None, None),
    ("", None, None),
    ("Source of Truth Hierarchy", SUBTITLE_FONT, None),
    ("Invoice > Admin Export > Employee List", None, None),
    ("", None, None),
    ("Sheet Descriptions", SUBTITLE_FONT, None),
    ("Invoice_Input     → Monthly invoice metadata: total, VAT, exchange rate, billing period", None, None),
    ("Employee_Master   → Cleaned employee data with division mappings", None, None),
    ("License_Raw       → Monthly Microsoft Admin Centre export (input)", None, None),
    ("User_Match        → Email-based mapping of licensed users to divisions", None, None),
    ("SKU_Cost_Input    → Manual input of invoice costs per SKU (yellow cells, excl VAT)", None, None),
    ("Allocation_Detail → Cost breakdown per user per SKU with formulas", None, None),
    ("Allocation_Pivot  → Clean allocation table: Division x SKU x Users x Cost", None, None),
    ("Division_Summary  → Final output for Finance with reconciliation check", None, None),
    ("Review_Queue      → Unresolved/flagged users requiring attention", None, None),
]

for i, (text, font, fill) in enumerate(readme_lines, 1):
    cell = ws.cell(row=i, column=1, value=text)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill

ws.column_dimensions["A"].width = 100
ws.sheet_view.showGridLines = False

# ===================================================
# Sheet 2: Invoice_Input
# ===================================================
print("  Building Invoice_Input...")
ws = wb.create_sheet("Invoice_Input")
ws.sheet_properties.tabColor = "C00000"

ws.cell(row=1, column=1, value="Monthly Invoice Input").font = TITLE_FONT
ws.cell(row=2, column=1, value="Complete ALL yellow cells before entering SKU amounts. This is the control sheet for reconciliation.").font = Font(italic=True, color="808080")

inv_labels = [
    ("Billing Period", "March 2026", "e.g. March 2026"),
    ("Invoice Number", "", "From Microsoft invoice"),
    ("Invoice Date", "", "Date on the invoice (DD/MM/YYYY)"),
    ("Snapshot Date", "", "Date the user list was exported from Admin Centre"),
    ("", "", ""),
    ("Invoice Total (incl VAT) ZAR", 0, "Exact Rand total from the invoice including VAT"),
    ("VAT Rate %", 0.15, "South Africa standard = 15%"),
    ("Invoice Total (excl VAT) ZAR", None, "Calculated: =incl / (1 + VAT rate)"),
    ("", "", ""),
    ("Exchange Rate (USD to ZAR)", 0, "Enter 0 if invoice is already in ZAR. Otherwise enter the rate used."),
    ("Invoice Total (USD)", 0, "Enter 0 if invoice is in ZAR. Otherwise enter USD total from invoice."),
    ("Converted Total (ZAR)", None, "Calculated: USD × rate. If invoice is already ZAR, this is ignored."),
    ("", "", ""),
    ("SKU Total (excl VAT) ZAR", None, "Auto-calculated from SKU_Cost_Input — must match Invoice Total excl VAT"),
    ("Variance (must be R0.00)", None, "= Invoice excl VAT − SKU total. RED if non-zero."),
    ("Reconciliation Status", None, "PASS or FAIL"),
]

for i, (label, val, note) in enumerate(inv_labels, 4):
    ws.cell(row=i, column=1, value=label).font = Font(bold=True) if label else Font()
    ws.cell(row=i, column=1).border = THIN_BORDER

    cell_b = ws.cell(row=i, column=2)
    cell_b.border = THIN_BORDER

    ws.cell(row=i, column=3, value=note).font = Font(italic=True, color="808080")

    if label == "":
        continue

    # Input cells (yellow)
    if label in ("Billing Period", "Invoice Number", "Invoice Date", "Snapshot Date",
                 "Invoice Total (incl VAT) ZAR", "VAT Rate %",
                 "Exchange Rate (USD to ZAR)", "Invoice Total (USD)"):
        cell_b.value = val
        cell_b.fill = INPUT_FILL
        if "ZAR" in label or "Total" in label:
            cell_b.number_format = CURRENCY_FMT
        if "Rate %" in label:
            cell_b.number_format = PCT_FMT

    # Calculated: invoice excl VAT = incl / (1 + VAT rate)
    if label == "Invoice Total (excl VAT) ZAR":
        cell_b.value = "=IF(B10=0,0,B9/(1+B10))"
        cell_b.number_format = CURRENCY_FMT
        cell_b.fill = KPI_FILL
    # Converted total from USD
    if label == "Converted Total (ZAR)":
        cell_b.value = "=IF(B13=0,0,B14*B13)"
        cell_b.number_format = CURRENCY_FMT
        cell_b.fill = KPI_FILL

# SKU total (excl VAT) — pulls from SKU_Cost_Input TOTAL row
# We'll set the actual formula after we know the SKU total row number.
# Store the row references for later.
inv_sku_total_row = 17   # row for "SKU Total (excl VAT) ZAR"
inv_variance_row = 18    # row for "Variance"
inv_status_row = 19      # row for "Reconciliation Status"

ws.cell(row=inv_sku_total_row, column=2).number_format = CURRENCY_FMT
ws.cell(row=inv_sku_total_row, column=2).fill = KPI_FILL

ws.cell(row=inv_variance_row, column=2).number_format = CURRENCY_FMT
ws.cell(row=inv_variance_row, column=2).font = Font(bold=True, size=12)

ws.cell(row=inv_status_row, column=2).font = Font(bold=True, size=14)

ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 55
ws.sheet_view.showGridLines = False

# ===================================================
# Sheet 3: Employee_Master
# ===================================================
print("  Building Employee_Master...")
ws = wb.create_sheet("Employee_Master")
ws.sheet_properties.tabColor = "4472C4"

emp_headers_out = ["Company Email", "Employee Name", "Employee Code", "Department",
                   "Department Group", "Site", "Division", "Job Title", "Supervisor",
                   "Cell Number", "Status"]
for col, h in enumerate(emp_headers_out, 1):
    ws.cell(row=1, column=col, value=h)
style_header_row(ws, 1, len(emp_headers_out))

for i, e in enumerate(employees, 2):
    ws.cell(row=i, column=1, value=e["email"])
    ws.cell(row=i, column=2, value=e["name"])
    ws.cell(row=i, column=3, value=e["code"])
    ws.cell(row=i, column=4, value=e["dept1"])
    ws.cell(row=i, column=5, value=e["dept2"])
    ws.cell(row=i, column=6, value=e["site"])
    ws.cell(row=i, column=7, value=e["division"])
    ws.cell(row=i, column=8, value=e["title"])
    ws.cell(row=i, column=9, value=e["supervisor"])
    ws.cell(row=i, column=10, value=e["cell"])
    ws.cell(row=i, column=11, value=e["status"])
    style_data_row(ws, i, len(emp_headers_out))

auto_width(ws, len(emp_headers_out))
freeze_and_filter(ws, "A2", len(emp_headers_out), 1)

# ===================================================
# Sheet 4: License_Raw
# ===================================================
print("  Building License_Raw...")
ws = wb.create_sheet("License_Raw")
ws.sheet_properties.tabColor = "4472C4"

raw_headers = ["Display Name", "User Principal Name", "Licenses"]
for col, h in enumerate(raw_headers, 1):
    ws.cell(row=1, column=col, value=h)
style_header_row(ws, 1, len(raw_headers))

for i, u in enumerate(licensed_users, 2):
    ws.cell(row=i, column=1, value=u["display_name"])
    ws.cell(row=i, column=2, value=u["email"])
    ws.cell(row=i, column=3, value="+".join(u["skus"]))
    style_data_row(ws, i, len(raw_headers))

auto_width(ws, len(raw_headers), 60)
freeze_and_filter(ws, "A2", len(raw_headers), 1)

# ===================================================
# Sheet 5: User_Match
# ===================================================
print("  Building User_Match...")
ws = wb.create_sheet("User_Match")
ws.sheet_properties.tabColor = "70AD47"

um_headers = ["Licensed Display Name", "Licensed Email", "Matched Employee Name",
              "Candidate Employee Email", "Employee Number", "Division",
              "Match Status", "Confidence", "Review Required", "User Status",
              "Mapping Note", "License SKU Count", "All Licenses Raw"]
for col, h in enumerate(um_headers, 1):
    ws.cell(row=1, column=col, value=h)
style_header_row(ws, 1, len(um_headers))

for i, m in enumerate(matched_users, 2):
    ws.cell(row=i, column=1, value=m["display_name"])
    ws.cell(row=i, column=2, value=m["email"])
    ws.cell(row=i, column=3, value=m["matched_name"])
    ws.cell(row=i, column=4, value=m["candidate_email"])
    ws.cell(row=i, column=5, value=m["emp_number"])
    ws.cell(row=i, column=6, value=m["division"])
    ws.cell(row=i, column=7, value=m["match_status"])
    ws.cell(row=i, column=8, value=m["confidence"])
    ws.cell(row=i, column=9, value=m["review_required"])
    ws.cell(row=i, column=10, value=m.get("user_status", "Active"))
    ws.cell(row=i, column=11, value=m["mapping_note"])
    ws.cell(row=i, column=12, value=len(m["skus"]))
    ws.cell(row=i, column=13, value="+".join(m["skus"]))
    style_data_row(ws, i, len(um_headers))

    # Highlight review rows
    if m["review_required"] == "Yes":
        for col in range(1, len(um_headers) + 1):
            ws.cell(row=i, column=col).fill = WARN_FILL
    # Highlight ex-employees in red
    elif m.get("user_status") == "Ex-Employee":
        for col in range(1, len(um_headers) + 1):
            ws.cell(row=i, column=col).fill = ERR_FILL

auto_width(ws, len(um_headers), 50)
freeze_and_filter(ws, "A2", len(um_headers), 1)

# ===================================================
# Sheet 6: SKU_Cost_Input
# ===================================================
print("  Building SKU_Cost_Input...")
ws = wb.create_sheet("SKU_Cost_Input")
ws.sheet_properties.tabColor = "ED7D31"

sku_headers = ["SKU", "Active Users", "Billing Hint",
               "Monthly Invoice Amount (excl VAT, ZAR)",
               "Unit Cost Per User (excl VAT)",
               "VAT Amount", "Total incl VAT",
               "Notes"]
for col, h in enumerate(sku_headers, 1):
    ws.cell(row=1, column=col, value=h)
style_header_row(ws, 1, len(sku_headers))

# Sort SKUs by user count descending
sorted_skus = sorted(all_skus, key=lambda s: sku_user_counts[s], reverse=True)

for i, sku in enumerate(sorted_skus, 2):
    ws.cell(row=i, column=1, value=sku)
    ws.cell(row=i, column=2, value=sku_user_counts[sku])
    ws.cell(row=i, column=3, value=billing_hint(sku))
    # Column D: input cell (yellow) - EXCL VAT amount from invoice line
    cell_d = ws.cell(row=i, column=4, value=0)
    cell_d.fill = INPUT_FILL
    cell_d.number_format = CURRENCY_FMT
    cell_d.font = Font(bold=True)
    # Column E: formula = D/B (unit cost excl VAT)
    ws.cell(row=i, column=5, value=f"=IF(B{i}=0,0,D{i}/B{i})")
    ws.cell(row=i, column=5).number_format = CURRENCY_FMT
    # Column F: VAT amount = D * VAT rate from Invoice_Input
    ws.cell(row=i, column=6, value=f"=D{i}*Invoice_Input!B10")
    ws.cell(row=i, column=6).number_format = CURRENCY_FMT
    # Column G: Total incl VAT = D + F
    ws.cell(row=i, column=7, value=f"=D{i}+F{i}")
    ws.cell(row=i, column=7).number_format = CURRENCY_FMT
    # Column H: note
    ws.cell(row=i, column=8, value="Enter the excl-VAT invoice amount for this SKU line item")
    style_data_row(ws, i, len(sku_headers))

# Total row
total_row = len(sorted_skus) + 2
ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row-1})").font = Font(bold=True)
ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row-1})")
ws.cell(row=total_row, column=4).font = Font(bold=True)
ws.cell(row=total_row, column=4).number_format = CURRENCY_FMT
ws.cell(row=total_row, column=5, value="").font = Font(bold=True)
ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row-1})")
ws.cell(row=total_row, column=6).font = Font(bold=True)
ws.cell(row=total_row, column=6).number_format = CURRENCY_FMT
ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{total_row-1})")
ws.cell(row=total_row, column=7).font = Font(bold=True)
ws.cell(row=total_row, column=7).number_format = CURRENCY_FMT
style_data_row(ws, total_row, len(sku_headers))
for col in range(1, len(sku_headers) + 1):
    ws.cell(row=total_row, column=col).fill = KPI_FILL

# Reconciliation check row
recon_row = total_row + 2
ws.cell(row=recon_row, column=1, value="RECONCILIATION CHECK").font = Font(bold=True, color="C00000")
ws.cell(row=recon_row, column=3, value="Invoice excl VAT").font = Font(bold=True)
ws.cell(row=recon_row, column=4, value="=Invoice_Input!B11")
ws.cell(row=recon_row, column=4).number_format = CURRENCY_FMT
ws.cell(row=recon_row + 1, column=3, value="SKU Total excl VAT").font = Font(bold=True)
ws.cell(row=recon_row + 1, column=4, value=f"=D{total_row}")
ws.cell(row=recon_row + 1, column=4).number_format = CURRENCY_FMT
ws.cell(row=recon_row + 2, column=3, value="Variance").font = Font(bold=True, color="C00000", size=12)
ws.cell(row=recon_row + 2, column=4, value=f"=D{recon_row}-D{recon_row+1}")
ws.cell(row=recon_row + 2, column=4).number_format = CURRENCY_FMT
ws.cell(row=recon_row + 2, column=4).font = Font(bold=True, color="C00000", size=12)
ws.cell(row=recon_row + 3, column=3, value="Status").font = Font(bold=True)
ws.cell(row=recon_row + 3, column=4, value=f'=IF(ABS(D{recon_row+2})<0.01,"PASS - Reconciled","FAIL - Check SKU amounts")')
ws.cell(row=recon_row + 3, column=4).font = Font(bold=True, size=12)

# Now wire up Invoice_Input sheet formulas that depend on SKU total row
inv_ws = wb["Invoice_Input"]
inv_ws.cell(row=inv_sku_total_row, column=2, value=f"=SKU_Cost_Input!D{total_row}")
inv_ws.cell(row=inv_variance_row, column=2, value=f"=B11-B{inv_sku_total_row}")
inv_ws.cell(row=inv_status_row, column=2, value=f'=IF(ABS(B{inv_variance_row})<0.01,"PASS","FAIL - SKU amounts do not match invoice")')

auto_width(ws, len(sku_headers), 50)
freeze_and_filter(ws, "A2", len(sku_headers), 1)

# ===================================================
# Sheet 7: Allocation_Detail
# ===================================================
print("  Building Allocation_Detail...")
ws = wb.create_sheet("Allocation_Detail")
ws.sheet_properties.tabColor = "4472C4"

ad_headers = ["Licensed Email", "Licensed Display Name", "Division", "Employee Number",
              "Match Status", "Review Required", "SKU", "Billing Hint",
              "Monthly Invoice Amount", "Active Users on SKU", "Unit Cost",
              "Allocated Cost", "Mapping Note"]
for col, h in enumerate(ad_headers, 1):
    ws.cell(row=1, column=col, value=h)
style_header_row(ws, 1, len(ad_headers))

# One row per user-SKU combination
sku_cost_count = len(sorted_skus)
detail_row = 2
for m in matched_users:
    for sku in m["skus"]:
        ws.cell(row=detail_row, column=1, value=m["email"])
        ws.cell(row=detail_row, column=2, value=m["display_name"])
        ws.cell(row=detail_row, column=3, value=m["division"])
        ws.cell(row=detail_row, column=4, value=m["emp_number"])
        ws.cell(row=detail_row, column=5, value=m["match_status"])
        ws.cell(row=detail_row, column=6, value=m["review_required"])
        ws.cell(row=detail_row, column=7, value=sku)
        ws.cell(row=detail_row, column=8, value=billing_hint(sku))
        # Column I: VLOOKUP to SKU_Cost_Input for invoice amount
        ws.cell(row=detail_row, column=9,
                value=f'=IFERROR(VLOOKUP(G{detail_row},SKU_Cost_Input!$A$2:$E${sku_cost_count+1},4,FALSE),0)')
        ws.cell(row=detail_row, column=9).number_format = CURRENCY_FMT
        # Column J: VLOOKUP to SKU_Cost_Input for active user count
        ws.cell(row=detail_row, column=10,
                value=f'=IFERROR(VLOOKUP(G{detail_row},SKU_Cost_Input!$A$2:$E${sku_cost_count+1},2,FALSE),0)')
        # Column K: Unit cost = Invoice / Users
        ws.cell(row=detail_row, column=11, value=f'=IF(J{detail_row}=0,0,I{detail_row}/J{detail_row})')
        ws.cell(row=detail_row, column=11).number_format = CURRENCY_FMT
        # Column L: Allocated cost = unit cost (1 user)
        ws.cell(row=detail_row, column=12, value=f'=K{detail_row}')
        ws.cell(row=detail_row, column=12).number_format = CURRENCY_FMT
        # Column M: Mapping note
        ws.cell(row=detail_row, column=13, value=m["mapping_note"])
        style_data_row(ws, detail_row, len(ad_headers))
        detail_row += 1

last_detail_row = detail_row - 1
print(f"  → {last_detail_row - 1} allocation detail rows")

auto_width(ws, len(ad_headers), 45)
freeze_and_filter(ws, "A2", len(ad_headers), 1)

# ===================================================
# Sheet 8: Allocation_Pivot (Clean Allocation Table)
# ===================================================
print("  Building Allocation_Pivot...")
ws = wb.create_sheet("Allocation_Pivot")
ws.sheet_properties.tabColor = "70AD47"

ws.cell(row=1, column=1, value="Clean Allocation Table: Division x SKU").font = TITLE_FONT
ws.cell(row=2, column=1, value="Cost values update automatically from SKU_Cost_Input").font = Font(italic=True, color="808080")

ap_headers = ["Division", "SKU", "Users", "Allocated Cost"]
for col, h in enumerate(ap_headers, 1):
    ws.cell(row=4, column=col, value=h)
style_header_row(ws, 4, len(ap_headers))

pivot_row = 5
for div in DIVISIONS:
    for sku in sorted_skus:
        ws.cell(row=pivot_row, column=1, value=div)
        ws.cell(row=pivot_row, column=2, value=sku)
        # Count users: COUNTIFS on Allocation_Detail
        ws.cell(row=pivot_row, column=3,
                value=f'=COUNTIFS(Allocation_Detail!$C$2:$C${last_detail_row},A{pivot_row},'
                      f'Allocation_Detail!$G$2:$G${last_detail_row},B{pivot_row})')
        # Sum cost: SUMIFS on Allocation_Detail
        ws.cell(row=pivot_row, column=4,
                value=f'=SUMIFS(Allocation_Detail!$L$2:$L${last_detail_row},'
                      f'Allocation_Detail!$C$2:$C${last_detail_row},A{pivot_row},'
                      f'Allocation_Detail!$G$2:$G${last_detail_row},B{pivot_row})')
        ws.cell(row=pivot_row, column=4).number_format = CURRENCY_FMT
        style_data_row(ws, pivot_row, len(ap_headers))
        if div in DIV_FILLS:
            ws.cell(row=pivot_row, column=1).fill = DIV_FILLS[div]
        pivot_row += 1

last_pivot_row = pivot_row - 1

# Grand total
ws.cell(row=pivot_row, column=1, value="TOTAL").font = Font(bold=True)
ws.cell(row=pivot_row, column=3, value=f'=SUM(C5:C{last_pivot_row})').font = Font(bold=True)
ws.cell(row=pivot_row, column=4, value=f'=SUM(D5:D{last_pivot_row})')
ws.cell(row=pivot_row, column=4).font = Font(bold=True)
ws.cell(row=pivot_row, column=4).number_format = CURRENCY_FMT
style_data_row(ws, pivot_row, len(ap_headers))
for col in range(1, len(ap_headers) + 1):
    ws.cell(row=pivot_row, column=col).fill = KPI_FILL

auto_width(ws, len(ap_headers), 50)
freeze_and_filter(ws, "A5", len(ap_headers), 4)

# ===================================================
# Sheet 9: Division_Summary (Finance Output)
# ===================================================
print("  Building Division_Summary...")
ws = wb.create_sheet("Division_Summary")
ws.sheet_properties.tabColor = "2F5496"
ws.sheet_view.showGridLines = False

# Title block
ws.cell(row=1, column=1, value="Microsoft License Allocation Summary").font = TITLE_FONT
ws.cell(row=2, column=1, value="March 2026").font = SUBTITLE_FONT
ws.cell(row=3, column=1, value="Enter monthly invoice amounts on SKU_Cost_Input. Values update automatically.").font = Font(italic=True, color="808080")

# KPI block
ws.cell(row=5, column=1, value="KPI").font = Font(bold=True)
ws.cell(row=5, column=2, value="Value").font = Font(bold=True)
for col in (1, 2):
    ws.cell(row=5, column=col).fill = HEADER_FILL
    ws.cell(row=5, column=col).font = HEADER_FONT
    ws.cell(row=5, column=col).border = THIN_BORDER

kpis = [
    ("Billing Period", "=Invoice_Input!B4", None),
    ("Licensed accounts in export", len(licensed_users), None),
    ("Matched to cost centre (High)", f'=COUNTIF(User_Match!$H$2:$H${len(matched_users)+1},"High")', None),
    ("Accounts needing review", f'=COUNTIF(User_Match!$I$2:$I${len(matched_users)+1},"Yes")', None),
    ("Invoice Total (excl VAT)", "=Invoice_Input!B11", CURRENCY_FMT),
    ("VAT Amount", f"=SKU_Cost_Input!F{len(sorted_skus)+2}", CURRENCY_FMT),
    ("Invoice Total (incl VAT)", "=Invoice_Input!B9", CURRENCY_FMT),
    ("Allocated Total (excl VAT)", f"=SKU_Cost_Input!D{len(sorted_skus)+2}", CURRENCY_FMT),
    ("Reconciliation Status", "=Invoice_Input!B19", None),
    ("Ex-Employee Licences (remove)", f'=COUNTIF(User_Match!$J$2:$J${len(matched_users)+1},"Ex-Employee")', None),
    ("Contractor Licences", f'=COUNTIF(User_Match!$J$2:$J${len(matched_users)+1},"Contractor")', None),
]
for i, (label, val, fmt) in enumerate(kpis, 6):
    ws.cell(row=i, column=1, value=label)
    ws.cell(row=i, column=2, value=val)
    ws.cell(row=i, column=1).fill = KPI_FILL
    ws.cell(row=i, column=2).fill = KPI_FILL
    ws.cell(row=i, column=1).border = THIN_BORDER
    ws.cell(row=i, column=2).border = THIN_BORDER
    if fmt:
        ws.cell(row=i, column=2).number_format = fmt
    if "reconciliation" in label.lower():
        ws.cell(row=i, column=2).font = Font(bold=True, size=12)

# Division summary table
div_table_row = 19
div_headers = ["Division", "Licensed Accounts", "Review Accounts",
               "SKU Assignments", "Allocated Monthly Cost (excl VAT)", "% of Total Cost"]
for col, h in enumerate(div_headers, 1):
    ws.cell(row=div_table_row, column=col, value=h)
style_header_row(ws, div_table_row, len(div_headers))

um_last = len(matched_users) + 1  # last row of User_Match data

for i, div in enumerate(DIVISIONS, div_table_row + 1):
    ws.cell(row=i, column=1, value=div)
    if div in DIV_FILLS:
        ws.cell(row=i, column=1).fill = DIV_FILLS[div]
    # Licensed accounts = COUNTIF User_Match Division
    ws.cell(row=i, column=2, value=f'=COUNTIF(User_Match!$F$2:$F${um_last},A{i})')
    # Review accounts
    ws.cell(row=i, column=3, value=f'=COUNTIFS(User_Match!$F$2:$F${um_last},A{i},User_Match!$I$2:$I${um_last},"Yes")')
    # SKU assignments from Allocation_Detail
    ws.cell(row=i, column=4, value=f'=COUNTIF(Allocation_Detail!$C$2:$C${last_detail_row},A{i})')
    # Allocated cost
    ws.cell(row=i, column=5, value=f'=SUMIF(Allocation_Detail!$C$2:$C${last_detail_row},A{i},Allocation_Detail!$L$2:$L${last_detail_row})')
    ws.cell(row=i, column=5).number_format = CURRENCY_FMT
    # % of total
    total_cost_cell = f"E{div_table_row + len(DIVISIONS) + 1}"
    ws.cell(row=i, column=6, value=f'=IF({total_cost_cell}=0,0,E{i}/{total_cost_cell})')
    ws.cell(row=i, column=6).number_format = PCT_FMT
    style_data_row(ws, i, len(div_headers))

# Total row
total_r = div_table_row + len(DIVISIONS) + 1
ws.cell(row=total_r, column=1, value="TOTAL").font = Font(bold=True)
ws.cell(row=total_r, column=2, value=f"=SUM(B{div_table_row+1}:B{total_r-1})").font = Font(bold=True)
ws.cell(row=total_r, column=3, value=f"=SUM(C{div_table_row+1}:C{total_r-1})").font = Font(bold=True)
ws.cell(row=total_r, column=4, value=f"=SUM(D{div_table_row+1}:D{total_r-1})").font = Font(bold=True)
ws.cell(row=total_r, column=5, value=f"=SUM(E{div_table_row+1}:E{total_r-1})")
ws.cell(row=total_r, column=5).font = Font(bold=True)
ws.cell(row=total_r, column=5).number_format = CURRENCY_FMT
ws.cell(row=total_r, column=6, value=f'=IF(E{total_r}=0,0,E{total_r}/E{total_r})')
ws.cell(row=total_r, column=6).number_format = PCT_FMT
ws.cell(row=total_r, column=6).font = Font(bold=True)
for col in range(1, len(div_headers) + 1):
    ws.cell(row=total_r, column=col).fill = KPI_FILL
    ws.cell(row=total_r, column=col).border = THIN_BORDER

# SKU breakdown section
sku_section_row = total_r + 3
ws.cell(row=sku_section_row, column=1, value="SKU Cost Breakdown").font = SUBTITLE_FONT
sku_table_row = sku_section_row + 1
sku_sum_headers = ["SKU", "Active Users", "Invoice Amount (excl VAT)", "Unit Cost Per User", "VAT Amount", "Total incl VAT"]
for col, h in enumerate(sku_sum_headers, 1):
    ws.cell(row=sku_table_row, column=col, value=h)
style_header_row(ws, sku_table_row, len(sku_sum_headers))

for i, sku in enumerate(sorted_skus, sku_table_row + 1):
    si = i - sku_table_row  # 1-based index into sorted_skus
    ws.cell(row=i, column=1, value=sku)
    ws.cell(row=i, column=2, value=f"=SKU_Cost_Input!B{si+1}")
    ws.cell(row=i, column=3, value=f"=SKU_Cost_Input!D{si+1}")
    ws.cell(row=i, column=3).number_format = CURRENCY_FMT
    ws.cell(row=i, column=4, value=f"=SKU_Cost_Input!E{si+1}")
    ws.cell(row=i, column=4).number_format = CURRENCY_FMT
    ws.cell(row=i, column=5, value=f"=SKU_Cost_Input!F{si+1}")
    ws.cell(row=i, column=5).number_format = CURRENCY_FMT
    ws.cell(row=i, column=6, value=f"=SKU_Cost_Input!G{si+1}")
    ws.cell(row=i, column=6).number_format = CURRENCY_FMT
    style_data_row(ws, i, len(sku_sum_headers))

# Audit trail note
audit_row = sku_table_row + len(sorted_skus) + 2
ws.cell(row=audit_row, column=1, value="Audit Trail & Controls").font = SUBTITLE_FONT
ws.cell(row=audit_row + 1, column=1, value="All costs trace: Microsoft Invoice (Invoice_Input) → SKU line items (SKU_Cost_Input) → Licensed Users (License_Raw) → Employee Mapping (User_Match) → Allocation (Allocation_Detail)")
ws.cell(row=audit_row + 2, column=1, value="Reconciliation: SKU excl-VAT total MUST equal invoice excl-VAT total. Any variance is flagged FAIL on Invoice_Input and above KPIs.")
ws.cell(row=audit_row + 3, column=1, value="VAT: Applied at total level via Invoice_Input rate. Individual SKU amounts are entered EXCLUDING VAT.")
ws.cell(row=audit_row + 4, column=1, value="Exchange Rate: If invoice is in USD, convert using the rate on Invoice_Input before entering SKU amounts in ZAR.")
ws.cell(row=audit_row + 5, column=1, value="Proration: Model uses a snapshot of users active at invoice date. Prorated amounts should be reflected in invoice line items.")
ws.cell(row=audit_row + 6, column=1, value="Unmatched users temporarily assigned to RS and flagged on Review_Queue for Finance confirmation.")

auto_width(ws, max(len(div_headers), len(sku_sum_headers)), 40)

# ===================================================
# Sheet 10: Review_Queue
# ===================================================
print("  Building Review_Queue...")
ws = wb.create_sheet("Review_Queue")
ws.sheet_properties.tabColor = "FFC000"

rq_headers = ["Licensed Display Name", "Licensed Email", "Assigned Division",
              "Matched Employee Name", "Candidate Employee Email", "Employee Number",
              "Confidence", "Review Type", "User Status", "Mapping Note", "Licenses Raw",
              "Action Required"]
for col, h in enumerate(rq_headers, 1):
    ws.cell(row=1, column=col, value=h)
style_header_row(ws, 1, len(rq_headers))

review_users = [m for m in matched_users if m["review_required"] == "Yes"]
for i, m in enumerate(review_users, 2):
    ws.cell(row=i, column=1, value=m["display_name"])
    ws.cell(row=i, column=2, value=m["email"])
    ws.cell(row=i, column=3, value=m["division"])
    ws.cell(row=i, column=4, value=m["matched_name"])
    ws.cell(row=i, column=5, value=m["candidate_email"])
    ws.cell(row=i, column=6, value=m["emp_number"])
    ws.cell(row=i, column=7, value=m["confidence"])
    ws.cell(row=i, column=8, value=m["match_status"])
    ws.cell(row=i, column=9, value=m.get("user_status", "Unknown"))
    ws.cell(row=i, column=10, value=m["mapping_note"])
    ws.cell(row=i, column=11, value="+".join(m["skus"]))
    ws.cell(row=i, column=12, value="Confirm division or update User_Match")
    style_data_row(ws, i, len(rq_headers))
    for col in range(1, len(rq_headers) + 1):
        ws.cell(row=i, column=col).fill = WARN_FILL

print(f"  → {len(review_users)} users in review queue")

auto_width(ws, len(rq_headers), 50)
freeze_and_filter(ws, "A2", len(rq_headers), 1)

# ===================================================
# Sheet 11: Licence_Cleanup (Ex-employees & flagged)
# ===================================================
print("  Building Licence_Cleanup...")
ws = wb.create_sheet("Licence_Cleanup")
ws.sheet_properties.tabColor = "C00000"

ws.cell(row=1, column=1, value="Licences Flagged for Removal / Management Review").font = TITLE_FONT
ws.cell(row=2, column=1, value="These users are no longer employees, are contractors, or need management review. Action: remove or reassign licences.").font = Font(italic=True, color="808080")

lc_headers = ["Display Name", "Email", "User Status", "Licences Held",
              "Licence Count", "Est. Monthly Cost (excl VAT)",
              "Reason / Note", "Action Required"]
for col, h in enumerate(lc_headers, 1):
    ws.cell(row=4, column=col, value=h)
style_header_row(ws, 4, len(lc_headers))

# Collect users to flag: ex-employees, contractors, management review
cleanup_users = [m for m in matched_users if m.get("user_status") in ("Ex-Employee", "Contractor", "Management Review")]

# Sort: ex-employees first, then contractors, then management review
status_order = {"Ex-Employee": 0, "Contractor": 1, "Management Review": 2}
cleanup_users.sort(key=lambda m: (status_order.get(m.get("user_status", ""), 99), m["display_name"]))

for i, m in enumerate(cleanup_users, 5):
    ws.cell(row=i, column=1, value=m["display_name"])
    ws.cell(row=i, column=2, value=m["email"])
    ws.cell(row=i, column=3, value=m.get("user_status", ""))
    ws.cell(row=i, column=4, value="+".join(m["skus"]))
    ws.cell(row=i, column=5, value=len(m["skus"]))
    # Estimated cost = sum of unit costs for each SKU (formula referencing Allocation_Detail)
    ws.cell(row=i, column=6,
            value=f'=SUMIF(Allocation_Detail!$A$2:$A${last_detail_row},B{i},Allocation_Detail!$L$2:$L${last_detail_row})')
    ws.cell(row=i, column=6).number_format = CURRENCY_FMT
    ws.cell(row=i, column=7, value=m["mapping_note"])

    # Set action based on status
    status = m.get("user_status", "")
    if status == "Ex-Employee":
        ws.cell(row=i, column=8, value="REMOVE LICENCE — no longer an employee")
        for col in range(1, len(lc_headers) + 1):
            ws.cell(row=i, column=col).fill = ERR_FILL
    elif status == "Contractor":
        ws.cell(row=i, column=8, value="Review: keep licence or remove (contractor)")
        for col in range(1, len(lc_headers) + 1):
            ws.cell(row=i, column=col).fill = WARN_FILL
    elif status == "Management Review":
        ws.cell(row=i, column=8, value="Escalate to management for decision")
        for col in range(1, len(lc_headers) + 1):
            ws.cell(row=i, column=col).fill = WARN_FILL

    style_data_row(ws, i, len(lc_headers))

last_cleanup_row = 4 + len(cleanup_users)

# Summary counts
summary_row = last_cleanup_row + 2
ws.cell(row=summary_row, column=1, value="Summary").font = SUBTITLE_FONT
ex_count = sum(1 for m in cleanup_users if m.get("user_status") == "Ex-Employee")
con_count = sum(1 for m in cleanup_users if m.get("user_status") == "Contractor")
mgmt_count = sum(1 for m in cleanup_users if m.get("user_status") == "Management Review")

summary_items = [
    ("Ex-Employees (remove licence)", ex_count, ERR_FILL),
    ("Contractors (review)", con_count, WARN_FILL),
    ("Management Review", mgmt_count, WARN_FILL),
    ("Total flagged accounts", len(cleanup_users), KPI_FILL),
]
for j, (label, count, fill) in enumerate(summary_items, summary_row + 1):
    ws.cell(row=j, column=1, value=label).font = Font(bold=True)
    ws.cell(row=j, column=2, value=count)
    ws.cell(row=j, column=1).fill = fill
    ws.cell(row=j, column=2).fill = fill
    ws.cell(row=j, column=1).border = THIN_BORDER
    ws.cell(row=j, column=2).border = THIN_BORDER

# Total potential savings
savings_row = summary_row + len(summary_items) + 2
ws.cell(row=savings_row, column=1, value="Potential Monthly Savings (excl VAT)").font = Font(bold=True, color="C00000", size=12)
ws.cell(row=savings_row, column=2, value=f"=SUM(F5:F{last_cleanup_row})")
ws.cell(row=savings_row, column=2).number_format = CURRENCY_FMT
ws.cell(row=savings_row, column=2).font = Font(bold=True, color="C00000", size=12)
ws.cell(row=savings_row, column=1).fill = KPI_FILL
ws.cell(row=savings_row, column=2).fill = KPI_FILL
ws.cell(row=savings_row, column=1).border = THIN_BORDER
ws.cell(row=savings_row, column=2).border = THIN_BORDER

print(f"  → {len(cleanup_users)} users flagged ({ex_count} ex-employees, {con_count} contractors, {mgmt_count} management review)")

auto_width(ws, len(lc_headers), 50)
freeze_and_filter(ws, "A5", len(lc_headers), 4)

# ===================================================
# Set sheet order & active sheet
# ===================================================
desired_order = ["Read_Me", "Invoice_Input", "Employee_Master", "License_Raw", "User_Match",
                 "SKU_Cost_Input", "Allocation_Detail", "Allocation_Pivot",
                 "Division_Summary", "Licence_Cleanup", "Review_Queue"]
current_names = wb.sheetnames
for target_idx, name in enumerate(desired_order):
    current_idx = wb.sheetnames.index(name)
    wb.move_sheet(name, offset=target_idx - current_idx)

# Set Division_Summary as active
wb.active = wb.sheetnames.index("Division_Summary")

# ===================================================
# Save
# ===================================================
print(f"\nSaving to: {OUTPUT}")
wb.save(OUTPUT)
print("Done! Workbook saved successfully.")

ex_emp_count = sum(1 for m in matched_users if m.get("user_status") == "Ex-Employee")
contractor_count = sum(1 for m in matched_users if m.get("user_status") == "Contractor")
print(f"\nSummary:")
print(f"  Employees loaded: {len(employees)}")
print(f"  Licensed users: {len(licensed_users)}")
print(f"  Unique SKUs: {len(all_skus)}")
print(f"  High confidence matches: {sum(1 for m in matched_users if m['confidence'] == 'High')}")
print(f"  Ex-employees (flag remove): {ex_emp_count}")
print(f"  Contractors: {contractor_count}")
print(f"  Review queue (remaining): {len(review_users)}")
print(f"  Licence cleanup total: {len(cleanup_users)}")
print(f"  Allocation detail rows: {last_detail_row - 1}")
