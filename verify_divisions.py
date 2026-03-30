"""Verify division assignments are correct by cross-checking User_Match vs Employee Master."""
import openpyxl

wb = openpyxl.load_workbook("Microsoft_License_Allocation_Model_March_2026.xlsx")
emp_wb = openpyxl.load_workbook("Source Data/employee_master_list.xlsx")

# --- Employee Master: build lookup by name ---
emp_sheet = emp_wb.active
emp_by_name = {}
emp_by_email = {}
for r in range(2, emp_sheet.max_row + 1):
    name = emp_sheet.cell(r, 1).value
    if not name:
        continue
    # Find division column - check headers
    division = None
    email = None
    emp_num = None
    for c in range(1, emp_sheet.max_column + 1):
        header = str(emp_sheet.cell(1, c).value or "").lower().strip()
        val = emp_sheet.cell(r, c).value
        if "division" in header or "department" in header:
            if val and str(val).strip() in ("RS", "AFS", "Namibia", "Mozambique"):
                division = str(val).strip()
        if "email" in header:
            email = str(val or "").lower().strip()
        if "employee" in header and "num" in header:
            emp_num = str(val or "")
    emp_by_name[str(name).strip().lower()] = {"division": division, "email": email, "emp_num": emp_num}
    if email:
        emp_by_email[email] = {"name": name, "division": division, "emp_num": emp_num}

# --- User_Match ---
um = wb["User_Match"]
users = []
for r in range(2, um.max_row + 1):
    name = um.cell(r, 1).value
    if not name:
        break
    users.append({
        "display_name": name,
        "email": (um.cell(r, 2).value or "").lower().strip(),
        "matched_name": um.cell(r, 3).value or "",
        "emp_number": um.cell(r, 5).value or "",
        "division": um.cell(r, 6).value or "",
        "match_status": um.cell(r, 7).value or "",
        "user_status": um.cell(r, 10).value or "",
        "mapping_note": um.cell(r, 11).value or "",
    })

# --- License_Raw ---
INV_UNIT = {
    "Microsoft 365 Business Standard": 209.55,
    "Microsoft 365 E3": 603.29,
    "Microsoft 365 Business Premium": 368.68,
    "Power BI Premium Per User": 402.19,
    "Exchange Online (Plan 1)": 67.03,
    "Power Automate per user plan": 251.37,
    "Microsoft Defender for Office 365 (Plan 2)": 83.79,
    "Power Apps per app plan (1 app or website)": 83.79,
}

lr = wb["License_Raw"]
license_map = {}
for r in range(2, lr.max_row + 1):
    if not lr.cell(r, 1).value:
        break
    email = (lr.cell(r, 2).value or "").lower()
    skus = []
    for sku in str(lr.cell(r, 3).value or "").split("+"):
        sku = sku.strip()
        if sku:
            skus.append(sku)
    license_map[email] = skus

# --- Check each user ---
print("=" * 80)
print("DIVISION ASSIGNMENT VERIFICATION")
print("=" * 80)

issues = []
div_detail = {"RS": [], "AFS": [], "Namibia": [], "Mozambique": []}

for u in users:
    skus = license_map.get(u["email"], [])
    paid_skus = [s for s in skus if INV_UNIT.get(s, 0) > 0]
    cost = sum(INV_UNIT.get(s, 0) for s in skus)

    # Check if employee master has a different division
    emp = emp_by_email.get(u["email"])
    master_div = emp["division"] if emp else None
    matched_name_lower = u["matched_name"].strip().lower()
    if not master_div and matched_name_lower in emp_by_name:
        master_div = emp_by_name[matched_name_lower]["division"]

    mismatch = False
    if master_div and master_div != u["division"]:
        mismatch = True
        issues.append({
            "name": u["display_name"],
            "email": u["email"],
            "model_div": u["division"],
            "master_div": master_div,
            "status": u["match_status"],
            "cost": cost,
        })

    if u["division"] in div_detail:
        div_detail[u["division"]].append({
            "name": u["display_name"],
            "cost": round(cost, 2),
            "paid_skus": len(paid_skus),
            "status": u["user_status"],
            "mismatch": mismatch,
            "master_div": master_div,
        })

# --- Print division summaries ---
for div in ["RS", "AFS", "Namibia", "Mozambique"]:
    entries = div_detail.get(div, [])
    total = round(sum(e["cost"] for e in entries), 2)
    active = sum(1 for e in entries if e["status"] == "Active")
    ex_emp = sum(1 for e in entries if e["status"] == "Ex-Employee")
    contr = sum(1 for e in entries if e["status"] == "Contractor")
    other = len(entries) - active - ex_emp - contr

    print(f"\n--- {div} ({len(entries)} users, R{total:,.2f}) ---")
    print(f"    Active: {active}  |  Ex-Employee: {ex_emp}  |  Contractor: {contr}  |  Other: {other}")
    for e in sorted(entries, key=lambda x: -x["cost"]):
        flag = " *** MASTER SAYS: " + str(e["master_div"]) + " ***" if e["mismatch"] else ""
        status_tag = f" [{e['status']}]" if e["status"] != "Active" else ""
        print(f"    {e['name']:<35} {e['paid_skus']} paid SKUs   R{e['cost']:>9,.2f}{status_tag}{flag}")

# --- Print issues ---
print("\n" + "=" * 80)
if issues:
    print(f"DIVISION MISMATCHES FOUND: {len(issues)}")
    for i in issues:
        print(f"  {i['name']}: Model={i['model_div']}, Master={i['master_div']} (Status: {i['status']}, Cost: R{i['cost']:,.2f})")
else:
    print("NO DIVISION MISMATCHES - All assignments match employee master")

# --- Final totals ---
print("\n" + "=" * 80)
print("DIVISION COST SUMMARY:")
grand = 0
for div in ["RS", "AFS", "Namibia", "Mozambique"]:
    entries = div_detail.get(div, [])
    total = round(sum(e["cost"] for e in entries), 2)
    grand += total
    print(f"  {div:<12}  {len(entries):>3} users  R{total:>10,.2f}")
print(f"  {'TOTAL':<12}  {sum(len(div_detail[d]) for d in div_detail):>3} users  R{grand:>10,.2f}")
print("=" * 80)
