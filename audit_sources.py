"""Source data audit — independent check of all inputs."""
import openpyxl, csv, os

BASE = r"c:\Users\nadhi\OneDrive - Wearcheck Reliability Solutions\Desktop\Microsoft Monthly Licensing"
SRC = os.path.join(BASE, "Source Data")

# 1. Employee master
emp_wb = openpyxl.load_workbook(os.path.join(SRC, "employee_master_list.xlsx"), read_only=True)
emp_ws = emp_wb["Employees Details (Onboarding a"]
headers = [str(c.value).strip() if c.value else "" for c in list(emp_ws.iter_rows(min_row=1, max_row=1))[0]]
print("EMP HEADERS:", headers[:15])

email_idx = next((i for i, h in enumerate(headers) if "e-mail" in h.lower()), None)
dept1_idx = next((i for i, h in enumerate(headers) if "department 1" in h.lower()), None)
dept2_idx = next((i for i, h in enumerate(headers) if "department 2" in h.lower()), None)
site_idx = next((i for i, h in enumerate(headers) if "site" in h.lower()), None)
fname_idx = next((i for i, h in enumerate(headers) if "first" in h.lower()), None)
sname_idx = next((i for i, h in enumerate(headers) if "surname" in h.lower() or "last" in h.lower()), None)
print(f"email={email_idx}, dept1={dept1_idx}, dept2={dept2_idx}, site={site_idx}, fname={fname_idx}, sname={sname_idx}")

divs = {"RS": 0, "AFS": 0, "Namibia": 0, "Mozambique": 0}
emp_emails = set()
total = 0
sample_afs = []
sample_nam = []
sample_moz = []
for row in emp_ws.iter_rows(min_row=2, values_only=True):
    email = str(row[email_idx]).strip() if email_idx is not None and email_idx < len(row) and row[email_idx] else ""
    if not email or email == "None":
        continue
    total += 1
    emp_emails.add(email.lower())
    d1 = str(row[dept1_idx]).strip() if dept1_idx and dept1_idx < len(row) and row[dept1_idx] else ""
    d2 = str(row[dept2_idx]).strip() if dept2_idx and dept2_idx < len(row) and row[dept2_idx] else ""
    site = str(row[site_idx]).strip() if site_idx and site_idx < len(row) and row[site_idx] else ""
    fname = str(row[fname_idx]).strip() if fname_idx and fname_idx < len(row) and row[fname_idx] else ""
    sname = str(row[sname_idx]).strip() if sname_idx and sname_idx < len(row) and row[sname_idx] else ""
    
    if d2 == "AFS" or d1 == "AFS":
        divs["AFS"] += 1
        if len(sample_afs) < 3:
            sample_afs.append(f"  {fname} {sname} ({email}) d1={d1} d2={d2}")
    elif "namibia" in d2.lower() or "namibia" in d1.lower() or "namibia" in site.lower():
        divs["Namibia"] += 1
        if len(sample_nam) < 3:
            sample_nam.append(f"  {fname} {sname} ({email}) d1={d1} d2={d2} site={site}")
    elif "mozambique" in d2.lower() or "mozambique" in d1.lower() or "mozambique" in site.lower() or "moz" in d2.lower():
        divs["Mozambique"] += 1
        if len(sample_moz) < 3:
            sample_moz.append(f"  {fname} {sname} ({email}) d1={d1} d2={d2} site={site}")
    else:
        divs["RS"] += 1
        
print(f"\nTotal employees with email: {total}")
print(f"Division breakdown: {divs}")
print(f"Sum check: {sum(divs.values())} == {total}")
print("Sample AFS:", sample_afs)
print("Sample Namibia:", sample_nam)
print("Sample Mozambique:", sample_moz)
emp_wb.close()

# 2. License_Raw
lic_wb = openpyxl.load_workbook(os.path.join(SRC, "Microsoft_License_Allocation_March_2026.xlsx"), read_only=True)
lr_ws = lic_wb["License_Raw"]
sku_counts = {}
users = set()
for row in lr_ws.iter_rows(min_row=2, values_only=True):
    upn = str(row[1]).strip() if row[1] else ""
    lic = str(row[2]).strip() if len(row) > 2 and row[2] else ""
    if not upn:
        continue
    users.add(upn.lower())
    for s in lic.split("+"):
        s = s.strip()
        if s:
            sku_counts[s] = sku_counts.get(s, 0) + 1
print(f"\nLicense_Raw: {len(users)} unique users")
print("SKU counts:")
for s in sorted(sku_counts.keys()):
    print(f"  {s}: {sku_counts[s]}")
lic_wb.close()

# 3. CSV
csv_total = 0
csv_lic = 0
csv_unlic = 0
csv_del = 0
csv_skus = {}
with open(os.path.join(SRC, "users_2026_03_30 11_05_42.csv"), "r", encoding="utf-8-sig") as f:
    for row in csv.DictReader(f, delimiter=";"):
        csv_total += 1
        deleted = row.get("Soft deletion time stamp", "").strip()
        if deleted:
            csv_del += 1
            continue
        lic = row.get("Licenses", "").strip()
        if lic == "Unlicensed" or not lic:
            csv_unlic += 1
            continue
        csv_lic += 1
        for s in lic.split("+"):
            s = s.strip()
            if s:
                csv_skus[s] = csv_skus.get(s, 0) + 1
print(f"\nCSV: {csv_total} total, {csv_lic} licensed, {csv_unlic} unlicensed, {csv_del} soft-deleted")

# 4. Cross check
paid = [
    ("Microsoft 365 Business Standard", 97),
    ("Microsoft 365 E3", 15),
    ("Microsoft 365 Business Premium", 23),
    ("Power BI Premium Per User", 7),
    ("Exchange Online (Plan 1)", 4),
    ("Power Automate per user plan", 1),
    ("Microsoft Defender for Office 365 (Plan 2)", 130),
    ("Power Apps per app plan (1 app or website)", 3),
]
inv_amounts = {
    "Microsoft 365 Business Standard": 20326.35,
    "Microsoft 365 E3": 9049.35,
    "Microsoft 365 Business Premium": 8479.64,
    "Power BI Premium Per User": 2815.33,
    "Exchange Online (Plan 1)": 268.12,
    "Power Automate per user plan": 251.37,
    "Microsoft Defender for Office 365 (Plan 2)": 10892.70,
    "Power Apps per app plan (1 app or website)": 251.37,
}

print("\nPAID SKU COMPARISON:")
print(f"{'SKU':<55} {'Admin':>6} {'CSV':>5} {'InvQty':>7} {'Match':>6} {'Delta':>6} {'Waste':>10}")
total_waste = 0
total_unassigned = 0
for sku, iq in paid:
    ac = sku_counts.get(sku, 0)
    cc = csv_skus.get(sku, 0)
    match = "YES" if ac == cc else "NO"
    delta = ac - iq
    unit = inv_amounts[sku] / iq
    unassigned = max(0, iq - ac)
    waste = unassigned * unit
    total_waste += waste
    total_unassigned += unassigned
    print(f"  {sku:<53} {ac:>6} {cc:>5} {iq:>7} {match:>6} {delta:>+6} R{waste:>9.2f}")
print(f"\n  TOTAL UNASSIGNED: {total_unassigned} licences")
print(f"  TOTAL WASTE: R{total_waste:.2f}/month excl VAT")

# 5. Check which licensed users have NO employee match
print("\n--- Licensed users NOT in employee master ---")
unmatched = 0
for row in openpyxl.load_workbook(os.path.join(SRC, "Microsoft_License_Allocation_March_2026.xlsx"), read_only=True)["License_Raw"].iter_rows(min_row=2, values_only=True):
    upn = str(row[1]).strip().lower() if row[1] else ""
    if upn and upn not in emp_emails:
        unmatched += 1
print(f"  {unmatched} licensed users have no direct email match in employee master")
print(f"  ({len(users)} total licensed, {len(emp_emails)} employees with email)")
