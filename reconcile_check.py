"""End-to-end reconciliation verification."""
import openpyxl

wb = openpyxl.load_workbook("Microsoft_License_Allocation_Model_March_2026.xlsx")

INV_UNIT = {
    "Microsoft 365 Business Standard": 209.55,
    "Microsoft 365 E3": 603.29,
    "Microsoft 365 Business Premium": 368.68,
    "Power BI Premium Per User": 402.19,
    "Exchange Online (Plan 1)": 67.03,
    "Power Automate per user plan": 251.37,
    "Microsoft Defender for Office 365 (Plan 2)": 83.79,
    "Power Apps per app plan (1 app or website)": 83.79,
}

sci = wb["SKU_Cost_Input"]
sku_data = {}
for r in range(2, 20):
    name = sci.cell(r, 1).value
    if not name or name == "TOTAL":
        break
    amt = float(sci.cell(r, 4).value or 0)
    users = int(sci.cell(r, 2).value or 0)
    inv_qty = sci.cell(r, 9).value
    sku_data[name] = {"amount": amt, "admin_users": users, "inv_qty": inv_qty}

invoice_total = sum(s["amount"] for s in sku_data.values())

lr = wb["License_Raw"]
license_map = {}
for r in range(2, lr.max_row + 1):
    if not lr.cell(r, 1).value:
        break
    email = (lr.cell(r, 2).value or "").lower()
    for sku in str(lr.cell(r, 3).value or "").split("+"):
        sku = sku.strip()
        if sku:
            license_map.setdefault(email, []).append(sku)

um = wb["User_Match"]
users = []
for r in range(2, um.max_row + 1):
    if not um.cell(r, 1).value:
        break
    users.append({
        "name": um.cell(r, 1).value,
        "email": (um.cell(r, 2).value or "").lower(),
        "division": um.cell(r, 6).value or "",
    })

sku_user_counts = {}
total_allocated = 0
for u in users:
    cost = 0
    for s in license_map.get(u["email"], []):
        cost += INV_UNIT.get(s, 0)
        sku_user_counts[s] = sku_user_counts.get(s, 0) + 1
    u["cost"] = round(cost, 2)
    total_allocated += u["cost"]

total_allocated = round(total_allocated, 2)
gap = round(invoice_total - total_allocated, 2)

print("=" * 70)
print("FULL RECONCILIATION CHECK  -  March 2026  -  INV-0303")
print("=" * 70)

print(f"\n1. TOTALS:")
print(f"   Invoice excl VAT:        R{invoice_total:>12,.2f}")
print(f"   Allocated to {len(users)} users:   R{total_allocated:>12,.2f}")
print(f"   Unallocated gap:         R{gap:>12,.2f}")

print(f"\n2. PER-SKU RECONCILIATION (paid SKUs only):")
header = f"   {'SKU':<48} {'InvAmt':>11} {'InvQty':>7} {'Users':>6} {'Alloc':>11} {'Unacc':>11}"
print(header)
print("   " + "-" * len(header))
total_accounted = 0
for name, d in sorted(sku_data.items(), key=lambda x: -x[1]["amount"]):
    if d["amount"] <= 0:
        continue
    actual = sku_user_counts.get(name, 0)
    alloc = round(actual * INV_UNIT.get(name, 0), 2)
    unacc = round(d["amount"] - alloc, 2)
    iq = str(d["inv_qty"]) if d["inv_qty"] not in (None, "N/A") else "-"
    total_accounted += alloc
    print(f"   {name:<48} R{d['amount']:>10,.2f} {iq:>7} {actual:>6} R{alloc:>10,.2f} R{unacc:>10,.2f}")

print(f"   {'TOTAL':<48} R{invoice_total:>10,.2f} {'':>7} {'':>6} R{round(total_accounted,2):>10,.2f} R{gap:>10,.2f}")

print(f"\n3. UNALLOCATED BREAKDOWN:")
gross_waste = 0
gross_saving = 0
for name, d in sorted(sku_data.items(), key=lambda x: -x[1]["amount"]):
    if d["amount"] <= 0:
        continue
    actual = sku_user_counts.get(name, 0)
    alloc = round(actual * INV_UNIT.get(name, 0), 2)
    unacc = round(d["amount"] - alloc, 2)
    iq = int(d["inv_qty"]) if d["inv_qty"] not in (None, "N/A") else 0
    unused = iq - actual
    if unacc > 0:
        gross_waste += unacc
        print(f"   WASTE:   {name}: R{unacc:>9,.2f}  ({unused} unused x R{INV_UNIT.get(name,0):,.2f})")
    elif unacc < 0:
        gross_saving += abs(unacc)
        print(f"   SAVING:  {name}: -R{abs(unacc):>8,.2f}  ({abs(unused)} extra users beyond billed qty)")
print(f"   ---")
print(f"   Gross waste:    R{gross_waste:>9,.2f}")
print(f"   Gross saving:  -R{gross_saving:>9,.2f}")
print(f"   Net unallocated: R{round(gross_waste - gross_saving, 2):>9,.2f}")

print(f"\n4. DIVISION TOTALS:")
div_costs = {}
div_users = {}
for u in users:
    div_costs[u["division"]] = round(div_costs.get(u["division"], 0) + u["cost"], 2)
    div_users[u["division"]] = div_users.get(u["division"], 0) + 1
for div in ["RS", "AFS", "Namibia", "Mozambique"]:
    if div in div_costs:
        pct = div_costs[div] / total_allocated * 100
        print(f"   {div:<12} {div_users[div]:>3} users   R{div_costs[div]:>10,.2f}   ({pct:.1f}%)")
print(f"   {'TOTAL':<12} {sum(div_users.values()):>3} users   R{sum(div_costs.values()):>10,.2f}   (100.0%)")

print(f"\n5. FINAL RECONCILIATION:")
print(f"   Invoice excl VAT:        R{invoice_total:>12,.2f}")
print(f"   - Allocated to users:    R{total_allocated:>12,.2f}")
print(f"   - Unallocated (waste):   R{gap:>12,.2f}")
print(f"   = Remainder:             R{round(invoice_total - total_allocated - gap, 2):>12,.2f}")
chk = "PASS" if round(invoice_total - total_allocated - gap, 2) == 0 else "FAIL"
print(f"   Reconciliation:          {chk}")
print("=" * 70)
