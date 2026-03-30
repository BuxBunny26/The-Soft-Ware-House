"""
WearCheck Microsoft License Allocation Dashboard
Flask web app — run with: python dashboard.py
Opens at http://localhost:5000
"""

from flask import Flask, render_template, jsonify, request, redirect, url_for, flash, send_from_directory
import openpyxl
import os
import json
import csv
import shutil
from datetime import datetime
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = "wck-license-2026"

BASE = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(BASE, "Microsoft_License_Allocation_Model_March_2026.xlsx")
MONTHLY_DIR = os.path.join(BASE, "Monthly_Data")
os.makedirs(MONTHLY_DIR, exist_ok=True)

ALLOWED_EXTENSIONS_INVOICE = {"pdf"}
ALLOWED_EXTENSIONS_USERS = {"csv", "xlsx"}

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
_cache = {}


def load_data():
    """Read all data from the output workbook."""
    wb = openpyxl.load_workbook(OUTPUT)
    data = {}

    # ---- User_Match ----
    um = wb["User_Match"]
    users = []
    for r in range(2, um.max_row + 1):
        name = um.cell(r, 1).value
        if not name:
            break
        users.append({
            "display_name": name,
            "email": um.cell(r, 2).value or "",
            "emp_name": um.cell(r, 3).value or "",
            "emp_email": um.cell(r, 4).value or "",
            "emp_number": um.cell(r, 5).value or "",
            "division": um.cell(r, 6).value or "",
            "match_status": um.cell(r, 7).value or "",
            "confidence": um.cell(r, 8).value or "",
            "review_required": um.cell(r, 9).value or "",
            "user_status": um.cell(r, 10).value or "",
            "mapping_note": um.cell(r, 11).value or "",
            "sku_count": um.cell(r, 12).value or 0,
            "all_licenses": str(um.cell(r, 13).value or ""),
            "duplicate_flag": um.cell(r, 14).value or "",
        })
    data["users"] = users

    # ---- SKU_Cost_Input (amounts are static) ----
    sci = wb["SKU_Cost_Input"]
    skus = []
    for r in range(2, 17):
        name = sci.cell(r, 1).value
        if not name or name == "TOTAL":
            break
        amt = sci.cell(r, 4).value
        amount = float(amt) if isinstance(amt, (int, float)) else 0.0
        active = sci.cell(r, 2).value or 0
        active = int(active) if isinstance(active, (int, float)) else 0
        inv_qty = sci.cell(r, 9).value
        qty_var = sci.cell(r, 10).value
        skus.append({
            "name": name,
            "active_users": active,
            "billing_hint": sci.cell(r, 3).value or "",
            "amount": amount,
            "unit_cost": round(amount / active, 2) if active > 0 else 0.0,
            "invoice_qty": inv_qty if inv_qty not in ("N/A", None) else None,
            "qty_variance": int(qty_var) if isinstance(qty_var, (int, float)) else None,
            "variance_note": sci.cell(r, 11).value or "",
        })
    data["skus"] = skus
    sku_lookup = {s["name"]: s for s in skus}
    data["sku_lookup"] = sku_lookup

    # ---- License_Raw (user -> SKU list; col 3 is '+'-delimited) ----
    lr = wb["License_Raw"]
    license_map = {}
    for r in range(2, lr.max_row + 1):
        display = lr.cell(r, 1).value
        email = lr.cell(r, 2).value
        sku_str = lr.cell(r, 3).value or ""
        if not display:
            break
        for sku in sku_str.split("+"):
            sku = sku.strip()
            if sku:
                license_map.setdefault((email or "").lower(), []).append(sku)
    data["license_map"] = license_map

    # ---- Invoice unit prices (list price per licence from INV-0303) ----
    INVOICE_UNIT_PRICES = {
        "Microsoft 365 Business Standard": 209.55,
        "Microsoft 365 E3": 603.29,
        "Microsoft 365 Business Premium": 368.68,
        "Power BI Premium Per User": 402.19,
        "Exchange Online (Plan 1)": 67.03,
        "Power Automate per user plan": 251.37,
        "Microsoft Defender for Office 365 (Plan 2)": 83.79,
        "Power Apps per app plan (1 app or website)": 83.79,
    }

    # Store invoice unit price on each SKU dict for templates
    for s in skus:
        s["invoice_unit"] = INVOICE_UNIT_PRICES.get(s["name"], 0.0)

    # ---- Compute per-user cost using invoice unit prices ----
    for u in users:
        email = u["email"].lower()
        user_skus = license_map.get(email, [])
        total_cost = 0.0
        for s in user_skus:
            total_cost += INVOICE_UNIT_PRICES.get(s, 0.0)
        u["allocated_cost"] = round(total_cost, 2)

    # ---- Build unallocated / waste breakdown per SKU ----
    # Count actual users per SKU from license_map (users who truly hold the licence)
    actual_sku_users = {}
    for email, email_skus in license_map.items():
        for s in email_skus:
            actual_sku_users[s] = actual_sku_users.get(s, 0) + 1

    waste_items = []
    total_allocated_to_users = sum(u["allocated_cost"] for u in users)
    for s in skus:
        if s["amount"] <= 0:
            continue
        inv_unit = INVOICE_UNIT_PRICES.get(s["name"], 0.0)
        inv_qty = s["invoice_qty"] if s["invoice_qty"] is not None else 0
        actual_users = actual_sku_users.get(s["name"], 0)
        allocated_for_sku = actual_users * inv_unit
        unallocated = round(s["amount"] - allocated_for_sku, 2)
        unused_licences = inv_qty - actual_users if inv_qty else 0
        if unallocated != 0 or unused_licences != 0:
            waste_items.append({
                "sku": s["name"],
                "invoice_total": s["amount"],
                "invoice_qty": inv_qty,
                "admin_users": actual_users,
                "invoice_unit": inv_unit,
                "allocated": round(allocated_for_sku, 2),
                "unallocated": unallocated,
                "unused_licences": unused_licences,
                "finding": (
                    f"{unused_licences} unused licence{'s' if abs(unused_licences) != 1 else ''} "
                    f"@ R{inv_unit:,.2f} = R{unallocated:,.2f}"
                    if unused_licences > 0
                    else (
                        f"{abs(unused_licences)} extra user{'s' if abs(unused_licences) != 1 else ''} "
                        f"beyond billed qty (saving R{abs(unallocated):,.2f})"
                        if unused_licences < 0
                        else f"COITE-managed, not in admin export"
                    )
                ),
            })
    data["waste_items"] = waste_items
    total_unallocated = round(sum(w["unallocated"] for w in waste_items if w["unallocated"] > 0), 2)
    total_savings = round(abs(sum(w["unallocated"] for w in waste_items if w["unallocated"] < 0)), 2)
    net_unallocated = round(sum(w["unallocated"] for w in waste_items), 2)
    data["total_unallocated"] = total_unallocated
    data["total_savings"] = total_savings
    data["net_unallocated"] = net_unallocated
    data["total_allocated_to_users"] = round(total_allocated_to_users, 2)

    # ---- Division aggregates ----
    division_costs = {}
    division_users = {}
    for u in users:
        div = u["division"]
        division_costs[div] = round(division_costs.get(div, 0) + u["allocated_cost"], 2)
        division_users[div] = division_users.get(div, 0) + 1
    data["division_costs"] = division_costs
    data["division_users"] = division_users

    # ---- Licence_Cleanup ----
    lc = wb["Licence_Cleanup"]
    cleanup = []
    for r in range(5, lc.max_row + 1):
        name = lc.cell(r, 1).value
        if not name or name in ("Summary", ""):
            break
        cleanup.append({
            "display_name": name,
            "email": lc.cell(r, 2).value or "",
            "user_status": lc.cell(r, 3).value or "",
            "licences": lc.cell(r, 4).value or "",
            "licence_count": lc.cell(r, 5).value or 0,
        })
    data["cleanup"] = cleanup

    # ---- COITE_Query ----
    cq = wb["COITE_Query"]
    queries = []
    for r in range(6, 12):
        num = cq.cell(r, 1).value
        sku = cq.cell(r, 2).value
        if not num:
            break
        unit_price = cq.cell(r, 6).value
        unassigned = cq.cell(r, 5).value
        try:
            waste = round(float(unit_price or 0) * int(unassigned or 0), 2)
        except (ValueError, TypeError):
            waste = 0
        queries.append({
            "number": num,
            "sku": sku or "",
            "invoice_qty": cq.cell(r, 3).value or "",
            "tenant_assigned": cq.cell(r, 4).value or "",
            "unassigned": unassigned or "",
            "unit_price": round(float(unit_price), 2) if isinstance(unit_price, (int, float)) else "N/A",
            "waste": waste,
        })
    data["queries"] = queries

    # ---- Review_Queue ----
    rq = wb["Review_Queue"]
    reviews = []
    for r in range(2, rq.max_row + 1):
        name = rq.cell(r, 1).value
        if not name or name == "DUPLICATE ACCOUNTS DETECTED":
            break
        reviews.append({
            "display_name": name,
            "email": rq.cell(r, 2).value or "",
            "division": rq.cell(r, 3).value or "",
            "confidence": rq.cell(r, 7).value or "",
            "review_type": rq.cell(r, 8).value or "",
            "user_status": rq.cell(r, 9).value or "",
            "note": rq.cell(r, 10).value or "",
            "licenses": rq.cell(r, 11).value or "",
            "action": rq.cell(r, 12).value or "",
        })

    # Duplicate accounts from Review_Queue
    duplicates = []
    in_dup_section = False
    for r in range(2, rq.max_row + 1):
        val = rq.cell(r, 1).value
        if val == "DUPLICATE ACCOUNTS DETECTED":
            in_dup_section = True
            continue
        if in_dup_section:
            if val == "Employee Number":
                continue  # header row
            if not val:
                break
            duplicates.append({
                "emp_number": val,
                "emp_name": rq.cell(r, 2).value or "",
                "email1": rq.cell(r, 3).value or "",
                "email2": rq.cell(r, 4).value or "",
                "email3": rq.cell(r, 5).value or "",
                "action": rq.cell(r, 6).value or "",
            })
    data["reviews"] = reviews
    data["duplicates"] = duplicates

    # ---- Admin_CSV_Snapshot summary ----
    acs = wb["Admin_CSV_Snapshot"]
    csv_summary = {}
    # Find the SKU summary section (look for "SKU Summary" header)
    csv_skus = []
    in_sku_section = False
    for r in range(1, acs.max_row + 1):
        val = acs.cell(r, 1).value
        if val == "SKU Summary: Admin Export vs CSV Snapshot":
            in_sku_section = True
            continue
        if in_sku_section:
            if val == "SKU":
                continue  # header row
            if not val:
                break
            csv_skus.append({
                "sku": val,
                "admin_count": acs.cell(r, 2).value or 0,
                "csv_count": acs.cell(r, 3).value or 0,
                "match": acs.cell(r, 4).value or "",
            })
    data["csv_skus"] = csv_skus

    # ---- KPIs (computed) ----
    invoice_total = sum(s["amount"] for s in skus)
    vat_rate = 0.15
    paid_skus = [s for s in skus if s["amount"] > 0]
    free_skus = [s for s in skus if s["amount"] == 0]
    discrepant = [s for s in skus if s["qty_variance"] is not None and s["qty_variance"] != 0]
    unassigned_total = sum(
        int(s["qty_variance"]) for s in skus
        if s["qty_variance"] is not None and isinstance(s["qty_variance"], (int, float)) and s["qty_variance"] > 0
    )

    data["kpis"] = {
        "total_licensed": len(users),
        "high_confidence": sum(1 for u in users if u["confidence"] == "High"),
        "medium_confidence": sum(1 for u in users if u["confidence"] == "Medium"),
        "low_confidence": sum(1 for u in users if u["confidence"] == "Low"),
        "review_required": sum(1 for u in users if u["review_required"] == "Yes"),
        "invoice_excl": round(invoice_total, 2),
        "invoice_vat": round(invoice_total * vat_rate, 2),
        "invoice_incl": round(invoice_total * (1 + vat_rate), 2),
        "ex_employees": sum(1 for u in users if u["user_status"] == "Ex-Employee"),
        "contractors": sum(1 for u in users if u["user_status"] == "Contractor"),
        "mgmt_review": sum(1 for u in users if u["user_status"] == "Management Review"),
        "active_users": sum(1 for u in users if u["user_status"] == "Active"),
        "paid_skus": len(paid_skus),
        "free_skus": len(free_skus),
        "total_skus": len(skus),
        "discrepant_skus": len(discrepant),
        "wasted_spend": total_unallocated,
        "wasted_spend_net": net_unallocated,
        "total_savings": total_savings,
        "total_allocated_to_users": round(total_allocated_to_users, 2),
        "coite_queries": len(queries),
        "cleanup_total": len(cleanup),
        "duplicate_groups": len(duplicates),
        "unassigned_licences": sum(w["unused_licences"] for w in waste_items if w["unused_licences"] > 0),
    }

    wb.close()
    return data


def get_data(force=False):
    global _cache
    if force or not _cache:
        _cache = load_data()
    return _cache


# ---------------------------------------------------------------------------
# Template filter
# ---------------------------------------------------------------------------
@app.template_filter("zar")
def zar_filter(value):
    try:
        return f"R{float(value):,.2f}"
    except (ValueError, TypeError):
        return "R0.00"


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.route("/")
def overview():
    data = get_data()
    return render_template("overview.html", data=data, active="overview")


@app.route("/users")
def users():
    data = get_data()
    return render_template("users.html", data=data, active="users")


@app.route("/allocation")
def allocation():
    data = get_data()
    return render_template("allocation.html", data=data, active="allocation")


@app.route("/actions")
def actions():
    data = get_data()
    return render_template("actions.html", data=data, active="actions")


@app.route("/api/reload")
def reload_data():
    get_data(force=True)
    return jsonify({"status": "ok", "message": "Data reloaded from workbook"})


@app.route("/invoice/main")
def serve_main_invoice():
    """Serve the main COITE invoice PDF from workspace root."""
    for f in os.listdir(BASE):
        if f.lower().endswith(".pdf") and "inv" in f.lower():
            return send_from_directory(BASE, f, as_attachment=False)
    flash("No invoice PDF found in project folder.", "danger")
    return redirect(url_for("overview"))


@app.route("/invoice/<folder>/<filename>")
def serve_monthly_invoice(folder, filename):
    """Serve a monthly snapshot invoice PDF."""
    safe_folder = secure_filename(folder)
    safe_file = secure_filename(filename)
    folder_path = os.path.join(MONTHLY_DIR, safe_folder)
    if not os.path.isfile(os.path.join(folder_path, safe_file)):
        flash("Invoice file not found.", "danger")
        return redirect(url_for("upload_page"))
    return send_from_directory(folder_path, safe_file, as_attachment=False)


# ---------------------------------------------------------------------------
# Upload helpers
# ---------------------------------------------------------------------------
def allowed_file(filename, allowed):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in allowed


def get_monthly_snapshots():
    """Return list of monthly snapshot folders sorted newest first."""
    snapshots = []
    if not os.path.exists(MONTHLY_DIR):
        return snapshots
    for folder in sorted(os.listdir(MONTHLY_DIR), reverse=True):
        folder_path = os.path.join(MONTHLY_DIR, folder)
        if not os.path.isdir(folder_path):
            continue
        meta_path = os.path.join(folder_path, "meta.json")
        if os.path.exists(meta_path):
            with open(meta_path, "r") as f:
                meta = json.load(f)
        else:
            meta = {"month": folder, "uploaded": "Unknown"}
        # Check what files exist
        files = os.listdir(folder_path)
        meta["has_invoice"] = any(f.endswith(".pdf") for f in files)
        meta["invoice_file"] = next((f for f in files if f.endswith(".pdf")), None)
        meta["has_users"] = any(f.endswith(".csv") or (f.endswith(".xlsx") and f != "meta.json") for f in files)
        meta["has_summary"] = os.path.exists(os.path.join(folder_path, "summary.json"))
        meta["folder"] = folder
        meta["files"] = [f for f in files if f != "meta.json" and f != "summary.json"]
        snapshots.append(meta)
    return snapshots


def parse_user_csv(filepath):
    """Parse a user CSV export and return list of user dicts.
    Auto-detects delimiter (comma or semicolon)."""
    with open(filepath, "r", encoding="utf-8-sig") as f:
        sample = f.read(2048)
        f.seek(0)
        # Detect delimiter
        if sample.count(";") > sample.count(","):
            delimiter = ";"
        else:
            delimiter = ","
        reader = csv.DictReader(f, delimiter=delimiter)
        users = []
        for row in reader:
            users.append(dict(row))
    return users


def build_month_summary(folder_path):
    """Build a summary.json from uploaded user data for comparison."""
    summary = {"users": [], "sku_counts": {}, "total_users": 0, "total_licensed": 0}

    # Common column name candidates
    NAME_COLS = ["Display name", "DisplayName", "User principal name"]
    LICENSE_COLS = ["Licenses", "Assigned licenses", "AssignedLicenses"]

    def find_col(row, candidates):
        for c in candidates:
            if c in row and row[c]:
                return row[c]
        return ""

    # Look for user CSV or XLSX
    for fname in os.listdir(folder_path):
        fpath = os.path.join(folder_path, fname)
        if fname.endswith(".csv") and not fname.startswith("meta"):
            rows = parse_user_csv(fpath)
            summary["total_users"] = len(rows)
            licensed_count = 0
            for row in rows:
                name = find_col(row, NAME_COLS)
                licenses = find_col(row, LICENSE_COLS)
                if not name:
                    continue
                # Filter out column-shift artifacts (semicolons in display names
                # cause DirSyncEnabled True/False to land in Licenses column)
                BOGUS_LICENSES = {"true", "false"}
                is_licensed = (licenses
                               and licenses.strip().lower() not in ("", "unlicensed")
                               and licenses.strip().lower() not in BOGUS_LICENSES)
                summary["users"].append({"name": name, "licenses": licenses})
                if is_licensed:
                    licensed_count += 1
                    for sku in str(licenses).split("+"):
                        sku = sku.strip()
                        if sku and sku.lower() not in BOGUS_LICENSES:
                            summary["sku_counts"][sku] = summary["sku_counts"].get(sku, 0) + 1
            summary["total_licensed"] = licensed_count
            break
        elif fname.endswith(".xlsx") and fname != "meta.json":
            wb = openpyxl.load_workbook(fpath, read_only=True)
            ws = wb.active
            headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1):
                vals = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
                name = ""
                for nc in NAME_COLS:
                    if nc in vals and vals[nc]:
                        name = str(vals[nc])
                        break
                if not name:
                    continue
                licenses = ""
                for lc in LICENSE_COLS:
                    if lc in vals and vals[lc]:
                        licenses = str(vals[lc])
                        break
                BOGUS_LICENSES = {"true", "false"}
                is_licensed = (licenses
                               and licenses.strip().lower() not in ("", "unlicensed")
                               and licenses.strip().lower() not in BOGUS_LICENSES)
                summary["users"].append({"name": name, "licenses": licenses})
                if is_licensed:
                    for sku in licenses.split("+"):
                        sku = sku.strip()
                        if sku and sku.lower() not in BOGUS_LICENSES:
                            summary["sku_counts"][sku] = summary["sku_counts"].get(sku, 0) + 1
            summary["total_users"] = len(summary["users"])
            summary["total_licensed"] = sum(1 for u in summary["users"]
                                           if u["licenses"] and u["licenses"].strip().lower() not in ("", "unlicensed"))
            wb.close()
            break

    with open(os.path.join(folder_path, "summary.json"), "w") as f:
        json.dump(summary, f, indent=2)
    return summary


def load_month_summary(folder):
    """Load summary.json for a given month folder."""
    path = os.path.join(MONTHLY_DIR, folder, "summary.json")
    if os.path.exists(path):
        with open(path, "r") as f:
            return json.load(f)
    return None


def compare_months(older_folder, newer_folder):
    """Compare two monthly snapshots and return differences."""
    older = load_month_summary(older_folder)
    newer = load_month_summary(newer_folder)
    if not older or not newer:
        return None

    # SKU count changes
    all_skus = set(list(older.get("sku_counts", {}).keys()) + list(newer.get("sku_counts", {}).keys()))
    sku_changes = []
    for sku in sorted(all_skus):
        old_count = older.get("sku_counts", {}).get(sku, 0)
        new_count = newer.get("sku_counts", {}).get(sku, 0)
        if old_count != new_count:
            sku_changes.append({
                "sku": sku,
                "old_count": old_count,
                "new_count": new_count,
                "change": new_count - old_count,
            })

    # User changes
    older_names = {u["name"] for u in older.get("users", [])}
    newer_names = {u["name"] for u in newer.get("users", [])}
    added = sorted(newer_names - older_names)
    removed = sorted(older_names - newer_names)

    return {
        "older_month": older_folder,
        "newer_month": newer_folder,
        "older_total": older.get("total_users", 0),
        "newer_total": newer.get("total_users", 0),
        "user_change": newer.get("total_users", 0) - older.get("total_users", 0),
        "sku_changes": sku_changes,
        "users_added": added,
        "users_removed": removed,
    }


# ---------------------------------------------------------------------------
# Upload & History Routes
# ---------------------------------------------------------------------------
@app.route("/upload", methods=["GET", "POST"])
def upload():
    if request.method == "POST":
        month_label = request.form.get("month", "").strip()
        if not month_label:
            flash("Please enter a month label (e.g. 2026-03).", "danger")
            return redirect(url_for("upload"))

        # Sanitise month label
        safe_month = secure_filename(month_label)
        folder_path = os.path.join(MONTHLY_DIR, safe_month)
        os.makedirs(folder_path, exist_ok=True)

        saved_files = []

        # Handle invoice PDF
        invoice_file = request.files.get("invoice")
        if invoice_file and invoice_file.filename and allowed_file(invoice_file.filename, ALLOWED_EXTENSIONS_INVOICE):
            fname = secure_filename(invoice_file.filename)
            invoice_file.save(os.path.join(folder_path, fname))
            saved_files.append(fname)

        # Handle user export CSV/XLSX
        users_file = request.files.get("users")
        if users_file and users_file.filename and allowed_file(users_file.filename, ALLOWED_EXTENSIONS_USERS):
            fname = secure_filename(users_file.filename)
            users_file.save(os.path.join(folder_path, fname))
            saved_files.append(fname)

        # Save metadata
        meta = {
            "month": month_label,
            "uploaded": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "files": saved_files,
        }
        with open(os.path.join(folder_path, "meta.json"), "w") as f:
            json.dump(meta, f, indent=2)

        # Build summary if user file uploaded
        if users_file and users_file.filename:
            try:
                build_month_summary(folder_path)
                flash(f"Uploaded {len(saved_files)} file(s) for {month_label} and built summary.", "success")
            except Exception as e:
                flash(f"Files saved but summary build failed: {e}", "warning")
        else:
            flash(f"Uploaded {len(saved_files)} file(s) for {month_label}.", "success")

        return redirect(url_for("upload"))

    snapshots = get_monthly_snapshots()
    return render_template("upload.html", snapshots=snapshots, active="upload")


@app.route("/upload/delete/<folder>", methods=["POST"])
def delete_snapshot(folder):
    safe = secure_filename(folder)
    folder_path = os.path.join(MONTHLY_DIR, safe)
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)
        flash(f"Deleted snapshot {safe}.", "info")
    return redirect(url_for("upload"))


@app.route("/compare")
def compare():
    snapshots = get_monthly_snapshots()
    # Only include snapshots that have summaries
    comparable = [s for s in snapshots if s.get("has_summary")]

    # Also include current month from the model
    current_data = get_data()
    current_summary = {
        "month": "Current (March 2026)",
        "folder": "__current__",
        "total_users": current_data["kpis"]["total_licensed"],
        "sku_counts": {},
    }
    for s in current_data["skus"]:
        current_summary["sku_counts"][s["name"]] = s["active_users"]

    older = request.args.get("older", "")
    newer = request.args.get("newer", "")
    comparison = None

    if older and newer:
        if newer == "__current__" or older == "__current__":
            # Compare against current model
            other = older if newer == "__current__" else newer
            other_summary = load_month_summary(other)
            if other_summary:
                all_skus = set(
                    list(other_summary.get("sku_counts", {}).keys()) +
                    list(current_summary["sku_counts"].keys())
                )
                sku_changes = []
                for sku in sorted(all_skus):
                    if older == "__current__":
                        old_c = current_summary["sku_counts"].get(sku, 0)
                        new_c = other_summary.get("sku_counts", {}).get(sku, 0)
                    else:
                        old_c = other_summary.get("sku_counts", {}).get(sku, 0)
                        new_c = current_summary["sku_counts"].get(sku, 0)
                    if old_c != new_c:
                        sku_changes.append({"sku": sku, "old_count": old_c, "new_count": new_c, "change": new_c - old_c})

                other_names = {u["name"] for u in other_summary.get("users", [])}
                current_names = {u["display_name"] for u in current_data["users"]}
                if older == "__current__":
                    comparison = {
                        "older_month": "Current (March 2026)", "newer_month": other,
                        "older_total": current_summary["total_users"],
                        "newer_total": other_summary.get("total_users", 0),
                        "user_change": other_summary.get("total_users", 0) - current_summary["total_users"],
                        "sku_changes": sku_changes,
                        "users_added": sorted(other_names - current_names),
                        "users_removed": sorted(current_names - other_names),
                    }
                else:
                    comparison = {
                        "older_month": other, "newer_month": "Current (March 2026)",
                        "older_total": other_summary.get("total_users", 0),
                        "newer_total": current_summary["total_users"],
                        "user_change": current_summary["total_users"] - other_summary.get("total_users", 0),
                        "sku_changes": sku_changes,
                        "users_added": sorted(current_names - other_names),
                        "users_removed": sorted(other_names - current_names),
                    }
        else:
            comparison = compare_months(older, newer)

    return render_template(
        "compare.html",
        snapshots=comparable,
        current=current_summary,
        comparison=comparison,
        selected_older=older,
        selected_newer=newer,
        active="compare",
    )


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("Loading workbook data...")
    get_data()
    print(f"Loaded {len(_cache['users'])} users, {len(_cache['skus'])} SKUs")
    port = int(os.environ.get("PORT", 5000))
    print(f"Starting dashboard at http://localhost:{port}")
    app.run(debug=False, host="0.0.0.0", port=port)
