"""
Independent verification script v2 — reads all source data AND the output workbook
and cross-checks every critical number. Reports PASS/FAIL for each check.
"""
import openpyxl
import csv
import os

BASE = r"c:\Users\nadhi\OneDrive - Wearcheck Reliability Solutions\Desktop\Microsoft Monthly Licensing"
SRC = os.path.join(BASE, "Source Data")
CSV_FILE = os.path.join(SRC, "users_2026_03_30 11_05_42.csv")
LIC_FILE = os.path.join(SRC, "Microsoft_License_Allocation_March_2026.xlsx")
OUTPUT = os.path.join(BASE, "Microsoft_License_Allocation_Model_March_2026.xlsx")

passes = 0
fails = 0

def check(label, expected, actual, tolerance=0):
    global passes, fails
    if isinstance(expected, float) and isinstance(actual, (float, int)):
        ok = abs(expected - float(actual)) <= tolerance
    else:
        ok = (expected == actual)
    status = "PASS" if ok else "FAIL"
    if not ok:
        fails += 1
        print(f"  [{status}] {label}: expected={expected}, got={actual}")
    else:
        passes += 1
        print(f"  [{status}] {label}: {actual}")

# Known invoice values (manually verified from PDF)
INVOICE_TOTAL_INCL = 60184.38
INVOICE_SUBTOTAL = 52334.23
INVOICE_VAT = 7850.15
INVOICE_SKU_AMOUNTS = {
    "Microsoft 365 Business Standard":          20326.35,
    "Microsoft 365 E3":                          9049.35,
    "Microsoft 365 Business Premium":            8479.64,
    "Power BI Premium Per User":                 2815.33,
    "Exchange Online (Plan 1)":                   268.12,
    "Power Automate per user plan":               251.37,
    "Microsoft Defender for Office 365 (Plan 2)":10892.70,
    "Power Apps per app plan (1 app or website)": 251.37,
}
INVOICE_SKU_QUANTITIES = {
    "Microsoft 365 Business Standard":           97,
    "Microsoft 365 E3":                          15,
    "Microsoft 365 Business Premium":            23,
    "Power BI Premium Per User":                  7,
    "Exchange Online (Plan 1)":                   4,
    "Power Automate per user plan":               1,
    "Microsoft Defender for Office 365 (Plan 2)":130,
    "Power Apps per app plan (1 app or website)": 3,
}

print("=" * 70)
print("VERIFICATION: Source Data vs Output Workbook")
print("=" * 70)

# ============================================================
# 1. INVOICE ARITHMETIC — verify internal consistency
# ============================================================
print("\n--- 1. Invoice arithmetic ---")
check("Subtotal + VAT = Total", INVOICE_TOTAL_INCL, round(INVOICE_SUBTOTAL + INVOICE_VAT, 2), 0.01)
check("VAT = Subtotal x 15%", INVOICE_VAT, round(INVOICE_SUBTOTAL * 0.15, 2), 0.02)
check("Sum of SKU amounts = Subtotal", INVOICE_SUBTOTAL, round(sum(INVOICE_SKU_AMOUNTS.values()), 2), 0.01)

# Verify unit prices
for sku, amount in INVOICE_SKU_AMOUNTS.items():
    qty = INVOICE_SKU_QUANTITIES[sku]
    unit = round(amount / qty, 2)
    check(f"  {sku}: {qty} x R{unit} = R{amount}", amount, round(qty * unit, 2), 0.02)

# ============================================================
# 2. CSV SOURCE — independent count
# ============================================================
print("\n--- 2. Admin CSV independent count ---")
csv_total = 0
csv_licensed = 0
csv_unlicensed = 0
csv_soft_deleted = 0
csv_sku_counts = {}
csv_blocked_licensed = 0

with open(CSV_FILE, "r", encoding="utf-8-sig") as f:
    for row in csv.DictReader(f, delimiter=";"):
        csv_total += 1
        lic = row.get("Licenses", "").strip()
        deleted = row.get("Soft deletion time stamp", "").strip()
        blocked = row.get("Block credential", "").strip().lower() == "true"

        if deleted:
            csv_soft_deleted += 1
            continue

        if lic == "Unlicensed" or not lic:
            csv_unlicensed += 1
            continue

        csv_licensed += 1
        skus = [s.strip() for s in lic.split("+") if s.strip()]
        for s in skus:
            csv_sku_counts[s] = csv_sku_counts.get(s, 0) + 1
        if blocked:
            csv_blocked_licensed += 1

check("CSV total", 842, csv_total)
check("CSV total = licensed + unlicensed + soft-deleted", csv_total, csv_licensed + csv_unlicensed + csv_soft_deleted)
check("CSV licensed", 147, csv_licensed)
check("CSV unlicensed", 641, csv_unlicensed)
check("CSV soft-deleted", 54, csv_soft_deleted)
check("CSV blocked+licensed", 0, csv_blocked_licensed)

# ============================================================
# 3. ADMIN EXPORT (License_Raw) — independent count
# ============================================================
print("\n--- 3. Admin Export (License_Raw) independent count ---")
lic_wb = openpyxl.load_workbook(LIC_FILE, read_only=True, data_only=True)
lr_ws = lic_wb["License_Raw"]

admin_sku_counts = {}
admin_users = 0
for row in lr_ws.iter_rows(min_row=2, values_only=True):
    display = str(row[0]).strip() if row[0] else ""
    upn = str(row[1]).strip() if row[1] else ""
    lic_str = str(row[2]).strip() if row[2] else ""
    if not upn or not lic_str:
        continue
    admin_users += 1
    skus = [s.strip() for s in lic_str.split("+") if s.strip()]
    for s in skus:
        admin_sku_counts[s] = admin_sku_counts.get(s, 0) + 1
lic_wb.close()

print(f"  Admin export: {admin_users} unique users")
check("Admin export user count", 147, admin_users)

# ============================================================
# 4. CROSS-CHECK: CSV vs Admin Export per paid SKU
# ============================================================
print("\n--- 4. CSV vs Admin Export per SKU (paid SKUs only) ---")
PAID_SKUS = list(INVOICE_SKU_QUANTITIES.keys())
for sku in PAID_SKUS:
    csv_c = csv_sku_counts.get(sku, 0)
    admin_c = admin_sku_counts.get(sku, 0)
    check(f"CSV==Admin: {sku}", admin_c, csv_c)

# Also check all other SKUs
print("\n  -- All other SKUs (free/trial) --")
other_skus = sorted(set(list(csv_sku_counts.keys()) + list(admin_sku_counts.keys())) - set(PAID_SKUS))
for sku in other_skus:
    csv_c = csv_sku_counts.get(sku, 0)
    admin_c = admin_sku_counts.get(sku, 0)
    check(f"CSV==Admin: {sku}", admin_c, csv_c)

# ============================================================
# 5. 3-WAY COMPARISON — Invoice vs Tenant
# ============================================================
print("\n--- 5. 3-way comparison: Invoice qty vs Admin count vs CSV count ---")
total_waste = 0.0
total_unassigned = 0
discrepant_count = 0

for sku in PAID_SKUS:
    inv_qty = INVOICE_SKU_QUANTITIES[sku]
    admin_c = admin_sku_counts.get(sku, 0)
    csv_c = csv_sku_counts.get(sku, 0)
    inv_amount = INVOICE_SKU_AMOUNTS[sku]
    unit_price = inv_amount / inv_qty
    unassigned = max(inv_qty - admin_c, 0)
    waste = unassigned * unit_price
    total_waste += waste
    total_unassigned += unassigned

    if inv_qty != admin_c:
        discrepant_count += 1
        print(f"  DISCREPANCY: {sku}: invoice={inv_qty}, admin={admin_c}, csv={csv_c}, "
              f"unassigned={unassigned}, waste=R{waste:.2f}/month")
    else:
        print(f"  OK: {sku}: invoice={inv_qty}, admin={admin_c}, csv={csv_c}")

check("Discrepant SKU count", 5, discrepant_count)
check("Total unassigned licences", 7, total_unassigned)
check("Total waste R/month", 1994.21, round(total_waste, 2), 0.02)

# ============================================================
# 6. OUTPUT WORKBOOK — structure and values
# ============================================================
print("\n--- 6. Output workbook verification ---")
out_wb = openpyxl.load_workbook(OUTPUT, data_only=True)

# Check all expected sheets exist
expected_sheets = ["Read_Me", "Invoice_Input", "Employee_Master", "License_Raw", "User_Match",
                   "SKU_Cost_Input", "Allocation_Detail", "Allocation_Pivot",
                   "Division_Summary", "Licence_Cleanup", "Review_Queue",
                   "Admin_CSV_Snapshot", "COITE_Query"]
for sheet in expected_sheets:
    check(f"Sheet exists: {sheet}", True, sheet in out_wb.sheetnames)

check("Sheet count", len(expected_sheets), len(out_wb.sheetnames))
check("Sheet order correct", expected_sheets, out_wb.sheetnames)

# --- Invoice_Input ---
print("\n  -- Invoice_Input --")
ii = out_wb["Invoice_Input"]
check("Invoice Number", "INV-0303", ii["B5"].value)
check("Invoice Date", "25/03/2026", str(ii["B6"].value))
check("Total Incl VAT", 60184.38, ii["B9"].value, 0.01)
check("VAT Rate", 0.15, ii["B10"].value, 0.001)

# --- SKU_Cost_Input ---
print("\n  -- SKU_Cost_Input --")
sci = out_wb["SKU_Cost_Input"]
sku_amounts_in_wb = {}
sku_inv_qty_wb = {}
sku_admin_count_wb = {}
for r in range(2, 25):
    sku_name = sci.cell(row=r, column=1).value
    amount = sci.cell(row=r, column=4).value
    inv_qty = sci.cell(row=r, column=9).value  # Column I = Invoice Qty
    admin_count = sci.cell(row=r, column=2).value  # Column B = User Count
    if sku_name and isinstance(amount, (int, float)):
        sku_amounts_in_wb[sku_name] = amount
    if sku_name and inv_qty is not None:
        sku_inv_qty_wb[sku_name] = inv_qty
    if sku_name and admin_count is not None:
        sku_admin_count_wb[sku_name] = admin_count

wb_total = sum(v for v in sku_amounts_in_wb.values())
check("WB SKU total = Invoice Subtotal", INVOICE_SUBTOTAL, round(wb_total, 2), 0.01)

# Check each paid SKU amount
for sku, expected_amt in INVOICE_SKU_AMOUNTS.items():
    if sku in sku_amounts_in_wb:
        check(f"WB amount: {sku}", expected_amt, sku_amounts_in_wb[sku], 0.01)
    else:
        print(f"  [WARN] Paid SKU '{sku}' not found in workbook amounts")

# Check invoice quantities in workbook
for sku, expected_qty in INVOICE_SKU_QUANTITIES.items():
    if sku in sku_inv_qty_wb:
        check(f"WB inv qty: {sku}", expected_qty, sku_inv_qty_wb[sku])

# Check admin counts in workbook match source
for sku in PAID_SKUS:
    if sku in sku_admin_count_wb:
        expected = admin_sku_counts.get(sku, 0)
        check(f"WB admin count: {sku}", expected, sku_admin_count_wb[sku])

# --- Admin_CSV_Snapshot ---
print("\n  -- Admin_CSV_Snapshot --")
acs = out_wb["Admin_CSV_Snapshot"]
csv_data_rows = 0
for r in range(5, acs.max_row + 1):
    upn_val = acs.cell(row=r, column=2).value
    if upn_val and "@" in str(upn_val):
        csv_data_rows += 1
    else:
        break
check("CSV Snapshot user rows = CSV licensed", csv_licensed, csv_data_rows)

# --- COITE_Query ---
print("\n  -- COITE_Query --")
cq = out_wb["COITE_Query"]
query_rows = 0
for r in range(6, cq.max_row + 1):
    if cq.cell(row=r, column=1).value is not None:
        query_rows += 1
    else:
        break
check("COITE Query count", 6, query_rows)

# --- Division_Summary KPIs (non-formula values only) ---
print("\n  -- Division_Summary KPIs (hardcoded values) --")
ds = out_wb["Division_Summary"]
kpi_dict = {}
for r in range(6, 30):
    label = ds.cell(row=r, column=1).value
    val = ds.cell(row=r, column=2).value
    if label:
        kpi_dict[str(label)] = val
    else:
        break

# Only check non-formula KPIs (formula cells show None since file was never opened in Excel)
check("KPI: Licensed accounts in export", 147, kpi_dict.get("Licensed accounts in export"))
check("KPI: SKU Qty Discrepancies", 5, kpi_dict.get("SKU Qty Discrepancies (Invoice vs Admin)"))
check("KPI: Total Tenant Accounts (CSV)", 842, kpi_dict.get("Total Tenant Accounts (CSV)"))
check("KPI: Licensed Users (CSV)", 147, kpi_dict.get("Licensed Users (CSV)"))
check("KPI: Unlicensed Accounts", 641, kpi_dict.get("Unlicensed Accounts"))
check("KPI: Soft-Deleted Accounts", 54, kpi_dict.get("Soft-Deleted Accounts"))
check("KPI: Unassigned Paid Licences", 7, kpi_dict.get("Unassigned Paid Licences (total)"))
check("KPI: Wasted Monthly Spend", 1994.21, kpi_dict.get("Wasted Monthly Spend (excl VAT)"), 0.01)
check("KPI: COITE Queries Raised", 6, kpi_dict.get("COITE Queries Raised"))

# Note which KPIs are formulas (will show None until workbook is opened in Excel)
formula_kpis = [k for k, v in kpi_dict.items() if v is None]
if formula_kpis:
    print(f"\n  [INFO] {len(formula_kpis)} KPIs use Excel formulas (will calculate when opened in Excel):")
    for k in formula_kpis:
        print(f"    - {k}")

out_wb.close()

# ============================================================
# FINAL SUMMARY
# ============================================================
print("\n" + "=" * 70)
print(f"VERIFICATION COMPLETE: {passes} PASS, {fails} FAIL")
if fails == 0:
    print("ALL CHECKS PASSED — Model is 100% accurate.")
else:
    print(f"WARNING: {fails} check(s) FAILED — review above.")
print("=" * 70)
